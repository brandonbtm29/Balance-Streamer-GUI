import serial
import serial.tools.list_ports
import time
import re
import csv
import datetime
import threading
import statistics

import tkinter as tk
from tkinter import filedialog, messagebox
import customtkinter as ctk

import matplotlib
matplotlib.use('TkAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

import os
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.chart import ScatterChart, Reference, Series
import scipy.stats
import scipy.signal

# --- Global Parsers ---
def parse_bonvoisin_response(response_str):
    match = re.search(r"([-+]?(?:\d+\.\d+|\d+))", response_str)
    if match:
        value = float(match.group(1))
        unit_match = re.search(r"([a-zA-Z]+)", response_str)
        unit = unit_match.group(1) if unit_match else "g"
        return "S", value, unit
    return None, None, None

mt_response_pattern = re.compile(r"S\s+([SD])\s+([ \-0-9\.]+)\s+([a-zA-Z]+)")
def parse_mt_sics_response(response_str):
    match = mt_response_pattern.search(response_str)
    if match:
        status, value_str, unit = match.groups()
        value = float(value_str.replace(' ', ''))
        return status, value, unit
    return None, None, None

ohaus_response_pattern = re.compile(r"([+-]?\s*[\d\.]+)\s+([a-zA-Z]+)")
def parse_ohaus_response(response_str):
    match = ohaus_response_pattern.search(response_str)
    if match:
        val_str, unit = match.groups()
        value = float(val_str.replace(' ', ''))
        return "S", value, unit
    return None, None, None


class BalanceTab(ctk.CTkFrame):
    def __init__(self, parent, root, tab_name, app=None):
        super().__init__(parent)
        self.root = root
        self.tab_name = tab_name
        self.app = app
        
        self.ser = None
        self.active_balance_brand = "Bonvoisin"
        self.baudrate = 9600
        
        # Data tracking
        self.times_sec = []
        self.times_min = []
        self.weights = []
        self.flow_rates = []
        self.timestamps = []
        self.start_time = None
        self.last_recorded_time = 0
        self.unsaved_changes = False
        
        self.is_running_thread = True
        self.recording = False
        
        self.current_unit = "g"
        self.unit_set = False
        self.smooth_window = 20
        self.smooth_type = 'mean'
        self.do_auto_scale_x = tk.BooleanVar(value=True)
        self.do_auto_scale_y_mass = tk.BooleanVar(value=True)
        self.do_auto_scale_y_flow = tk.BooleanVar(value=True)
        self.show_flow = tk.BooleanVar(value=True)
        self.do_auto_stop = tk.BooleanVar(value=False)
        self.last_activity_time = time.time()
        
        # Backup file
        # Ensure backup directory exists
        backup_dir = os.path.join(os.getcwd(), "Data Backups")
        os.makedirs(backup_dir, exist_ok=True)
        self.backup_path = os.path.join(backup_dir, f"{self.tab_name.replace(' ', '_')}_Backup.csv")
        self.prune_backup_file()
        
        # Construct UI
        self.build_ui()
        
        # Load User Settings
        self.load_tab_settings()
        
        # Start Threads
        self.serial_thread = threading.Thread(target=self.serial_worker, daemon=True)
        self.serial_thread.start()

    def load_tab_settings(self):
        if not self.app: return
        kb = self.app.config.get("known_balances", {})
        if self.tab_name in kb:
            settings = kb[self.tab_name].get("settings", {})
            if "interval" in settings:
                self.ent_interval.delete(0, tk.END)
                self.ent_interval.insert(0, settings["interval"])
            if "savgol_win" in settings:
                self.ent_savgol_win.delete(0, tk.END)
                self.ent_savgol_win.insert(0, settings["savgol_win"])
            if "savgol_poly" in settings:
                self.ent_savgol_poly.delete(0, tk.END)
                self.ent_savgol_poly.insert(0, settings["savgol_poly"])
            if "filter_type" in settings:
                self.combo_filter.set(settings["filter_type"])
                self.on_filter_change(settings["filter_type"])
            if "filter_param" in settings:
                self.ent_filter_param.delete(0, tk.END)
                self.ent_filter_param.insert(0, settings["filter_param"])
            if "rpm" in settings:
                self.ent_rpm.delete(0, tk.END)
                self.ent_rpm.insert(0, settings["rpm"])
            if "rollers" in settings:
                self.ent_rollers.delete(0, tk.END)
                self.ent_rollers.insert(0, settings["rollers"])
            if "do_auto_scale_x" in settings:
                self.do_auto_scale_x.set(settings["do_auto_scale_x"])
            if "do_auto_scale_y_mass" in settings:
                self.do_auto_scale_y_mass.set(settings["do_auto_scale_y_mass"])
            if "do_auto_scale_y_flow" in settings:
                self.do_auto_scale_y_flow.set(settings["do_auto_scale_y_flow"])
            if "do_auto_stop" in settings:
                self.do_auto_stop.set(settings["do_auto_stop"])
            if "auto_stop_min" in settings and hasattr(self, 'ent_auto_stop_min'):
                self.ent_auto_stop_min.delete(0, tk.END)
                self.ent_auto_stop_min.insert(0, settings["auto_stop_min"])
            if "auto_stop_thresh" in settings and hasattr(self, 'ent_auto_stop_thresh'):
                self.ent_auto_stop_thresh.delete(0, tk.END)
                self.ent_auto_stop_thresh.insert(0, settings["auto_stop_thresh"])
            if "experiment_notes" in settings and hasattr(self, 'txt_notes'):
                self.txt_notes.delete("0.0", tk.END)
                self.txt_notes.insert(tk.END, settings["experiment_notes"])

    def save_tab_settings(self):
        if not self.app: return
        kb = self.app.config.get("known_balances", {})
        if self.tab_name not in kb:
            kb[self.tab_name] = {"brand": self.combo_brand.get(), "port": self.combo_com.get(), "unsaved": False}
        
        kb[self.tab_name]["settings"] = {
            "interval": self.ent_interval.get(),
            "savgol_win": self.ent_savgol_win.get(),
            "savgol_poly": self.ent_savgol_poly.get(),
            "filter_type": self.combo_filter.get(),
            "filter_param": self.ent_filter_param.get(),
            "rpm": self.ent_rpm.get(),
            "rollers": self.ent_rollers.get(),
            "do_auto_scale_x": self.do_auto_scale_x.get(),
            "do_auto_scale_y_mass": self.do_auto_scale_y_mass.get(),
            "do_auto_scale_y_flow": self.do_auto_scale_y_flow.get(),
            "do_auto_stop": getattr(self, 'do_auto_stop', tk.BooleanVar()).get(),
            "auto_stop_min": getattr(self, 'ent_auto_stop_min', ctk.CTkEntry(self)).get(),
            "auto_stop_thresh": getattr(self, 'ent_auto_stop_thresh', ctk.CTkEntry(self)).get(),
            "experiment_notes": getattr(self, 'txt_notes', ctk.CTkTextbox(self)).get("0.0", tk.END).strip()
        }

    def prune_backup_file(self):
        if not os.path.exists(self.backup_path):
            return
            
        try:
            with open(self.backup_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                rows = list(reader)
                
            if not header:
                return
                
            valid_rows = []
            thirty_days_ago = datetime.datetime.now() - datetime.timedelta(days=30)
            
            for row in rows:
                if len(row) > 0:
                    try:
                        ts_str = row[0]
                        if len(ts_str) <= 19:
                            ts = datetime.datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                        else:
                            ts = datetime.datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S.%f")
                        if ts >= thirty_days_ago:
                            valid_rows.append(row)
                    except:
                        valid_rows.append(row)
                        
            with open(self.backup_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(valid_rows)
        except Exception as e:
            print(f"[{self.tab_name}] Error pruning backup: {e}")

    def build_ui(self):
        self.main_paned = tk.PanedWindow(self, orient=tk.HORIZONTAL, bg="#2b2b2b", sashwidth=6, sashrelief=tk.RAISED)
        self.main_paned.pack(fill="both", expand=True, padx=5, pady=5)
        
        self.wrapper_left = ctk.CTkFrame(self.main_paned, width=320)
        self.frame_left = ctk.CTkScrollableFrame(self.wrapper_left, width=320)
        self.frame_left.pack(fill="both", expand=True)
        
        self.frame_mid = ctk.CTkFrame(self.main_paned)
        
        self.wrapper_right = ctk.CTkFrame(self.main_paned, width=320)
        self.frame_right = ctk.CTkScrollableFrame(self.wrapper_right, width=320)
        self.frame_right.pack(fill="both", expand=True)
        
        self.main_paned.add(self.wrapper_left, minsize=200)
        self.main_paned.add(self.frame_mid, minsize=400, stretch="always")
        self.main_paned.add(self.wrapper_right, minsize=200)
        
        def make_entry_row(parent, label_text, default_val):
            f = ctk.CTkFrame(parent, fg_color="transparent")
            f.pack(fill="x", padx=10, pady=2)
            ctk.CTkLabel(f, text=label_text).pack(side="left")
            ent = ctk.CTkEntry(f, width=80)
            ent.insert(0, default_val)
            ent.pack(side="right")
            return ent
            
        # --- LEFT PANEL: Recording, Filtering, Calibration, Axis ---
        ctk.CTkLabel(self.frame_left, text="Recording Settings", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=(5, 5), padx=10)
        self.ent_interval = make_entry_row(self.frame_left, "Interval (s):", "1.0")
        
        ctk.CTkCheckBox(self.frame_left, text="Enable Auto-Stop", variable=self.do_auto_stop).pack(anchor="w", padx=10, pady=(5, 2))
        self.ent_auto_stop_min = make_entry_row(self.frame_left, "Auto-Stop After (min):", "5.0")
        self.ent_auto_stop_thresh = make_entry_row(self.frame_left, "Flow Threshold:", "0.1")
        
        self.btn_record = ctk.CTkButton(self.frame_left, text="Start Recording", fg_color="#2ecc71", hover_color="#27ae60", command=self.toggle_record)
        self.btn_record.pack(fill="x", padx=10, pady=(10, 5))
        
        ctk.CTkFrame(self.frame_left, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_left, text="Derivative Engine", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10)
        self.ent_savgol_win = make_entry_row(self.frame_left, "SavGol Window:", "5")
        self.ent_savgol_poly = make_entry_row(self.frame_left, "SavGol Poly:", "3")
        
        ctk.CTkFrame(self.frame_left, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_left, text="Secondary Filtering", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10)
        self.combo_filter = ctk.CTkComboBox(self.frame_left, values=["Mean", "Median", "EMA", "Butterworth", "Adaptive"], command=self.on_filter_change)
        self.combo_filter.set("Mean")
        self.combo_filter.pack(fill="x", padx=10, pady=5)
        
        self.lbl_filter_param = ctk.CTkLabel(self.frame_left, text="Window Size (samples):")
        self.lbl_filter_param.pack(anchor="w", padx=10)
        self.ent_filter_param = ctk.CTkEntry(self.frame_left)
        self.ent_filter_param.insert(0, "20")
        self.ent_filter_param.pack(fill="x", padx=10, pady=2)
        
        ctk.CTkFrame(self.frame_left, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_left, text="Pump Calibration", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10)
        self.ent_rpm = make_entry_row(self.frame_left, "Pump RPM:", "30")
        self.ent_rollers = make_entry_row(self.frame_left, "Rollers:", "3")
        self.btn_log_cal = ctk.CTkButton(self.frame_left, text="Log Calibration Point", fg_color="#8e44ad", hover_color="#732d91", command=self.log_calibration)
        self.btn_log_cal.pack(fill="x", padx=10, pady=(5, 10))
        
        ctk.CTkFrame(self.frame_left, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_left, text="Axis Limits", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10)
        
        f_auto = ctk.CTkFrame(self.frame_left, fg_color="transparent")
        f_auto.pack(fill="x", padx=10, pady=2)
        ctk.CTkCheckBox(f_auto, text="Auto Time", variable=self.do_auto_scale_x).pack(side="left", padx=2)
        
        f_auto2 = ctk.CTkFrame(self.frame_left, fg_color="transparent")
        f_auto2.pack(fill="x", padx=10, pady=2)
        ctk.CTkCheckBox(f_auto2, text="Auto Mass", variable=self.do_auto_scale_y_mass).pack(side="left", padx=2)
        ctk.CTkCheckBox(f_auto2, text="Auto Flow", variable=self.do_auto_scale_y_flow).pack(side="left", padx=10)
        
        ctk.CTkLabel(self.frame_left, text="Plot X-Axis By:", font=ctk.CTkFont(size=12)).pack(anchor="w", padx=10, pady=(5,0))
        self.var_xaxis = ctk.StringVar(value="Duration")
        self.seg_xaxis = ctk.CTkSegmentedButton(self.frame_left, values=["Duration", "Timestamp"], variable=self.var_xaxis)
        self.seg_xaxis.pack(fill="x", padx=10, pady=(0, 10))
        
        self.ent_xmin = make_entry_row(self.frame_left, "X Min (min):", "0")
        self.ent_xmax = make_entry_row(self.frame_left, "X Max (min):", "10")
        self.ent_ymin = make_entry_row(self.frame_left, "Mass Min (g):", "0")
        self.ent_ymax = make_entry_row(self.frame_left, "Mass Max (g):", "100")
        self.ent_flowmin = make_entry_row(self.frame_left, "Flow Min:", "0")
        self.ent_flowmax = make_entry_row(self.frame_left, "Flow Max:", "20")
        
        self.btn_apply_lims = ctk.CTkButton(self.frame_left, text="Apply Manual Limits", command=self.apply_axis_limits)
        self.btn_apply_lims.pack(fill="x", padx=10, pady=(10, 10))
        
        ctk.CTkFrame(self.frame_left, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_left, text="Data Analysis", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10)
        self.ent_fit_start = make_entry_row(self.frame_left, "Fit Start (min):", "0")
        self.ent_fit_end = make_entry_row(self.frame_left, "Fit End (min):", "10")
        
        self.btn_apply_fit = ctk.CTkButton(self.frame_left, text="Apply Linear Fit", fg_color="#e67e22", hover_color="#d35400", command=self.apply_linear_fit)
        self.btn_apply_fit.pack(fill="x", padx=10, pady=(5, 10))
        
        # --- RIGHT PANEL: Connection, Notes & Export ---
        ctk.CTkLabel(self.frame_right, text="Connection", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=(5, 5), padx=10)
        
        ctk.CTkLabel(self.frame_right, text="Balance Brand:", font=ctk.CTkFont(size=12)).pack(anchor="w", padx=10, pady=(0, 0))
        self.combo_brand = ctk.CTkComboBox(self.frame_right, values=["Bonvoisin", "Mettler Toledo", "Ohaus Adventurer", "Lachoi"])
        self.combo_brand.set("Bonvoisin")
        self.combo_brand.pack(fill="x", padx=10, pady=(2, 10))
        
        ctk.CTkLabel(self.frame_right, text="COM Port:").pack(anchor="w", padx=10)
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.combo_com = ctk.CTkComboBox(self.frame_right, values=ports if ports else ["No Ports Found"])
        if ports: self.combo_com.set(ports[0])
        self.combo_com.pack(fill="x", padx=10, pady=5)
        
        self.btn_refresh = ctk.CTkButton(self.frame_right, text="Refresh Ports", command=self.refresh_com_ports)
        self.btn_refresh.pack(fill="x", padx=10, pady=5)
        
        self.btn_connect = ctk.CTkButton(self.frame_right, text="Connect", fg_color="#3498db", hover_color="#2980b9", command=self.connect_serial)
        self.btn_connect.pack(fill="x", padx=10, pady=(10, 2))
        
        self.btn_disconnect = ctk.CTkButton(self.frame_right, text="Disconnect", fg_color="#e74c3c", hover_color="#c0392b", command=self.disconnect_serial)
        self.btn_disconnect.pack(fill="x", padx=10, pady=(2, 5))
        
        self.lbl_status = ctk.CTkLabel(self.frame_right, text="Status: Disconnected", text_color="gray", font=ctk.CTkFont(weight="bold"))
        self.lbl_status.pack(anchor="center", pady=(0, 10))
        
        ctk.CTkFrame(self.frame_right, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.frame_right, text="Experiment Notes", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=10, pady=(5, 5))
        self.txt_notes = ctk.CTkTextbox(self.frame_right, height=150)
        self.txt_notes.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkFrame(self.frame_right, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        self.btn_tare = ctk.CTkButton(self.frame_right, text="Tare Balance", command=self.tare_balance)
        self.btn_tare.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkFrame(self.frame_right, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        self.btn_save_excel = ctk.CTkButton(self.frame_right, text="Save to Excel", fg_color="#27ae60", hover_color="#219150", font=ctk.CTkFont(weight="bold"), command=self.save_excel)
        self.btn_save_excel.pack(fill="x", padx=10, pady=5)
        
        self.btn_save_graph = ctk.CTkButton(self.frame_right, text="Save Graph PNG", fg_color="#2980b9", hover_color="#1f618d", command=self.save_graph)
        self.btn_save_graph.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkFrame(self.frame_right, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        
        self.btn_recover = ctk.CTkButton(self.frame_right, text="Recover Session", fg_color="#8e44ad", hover_color="#732d91", command=lambda: self.app.prompt_session_recovery(self.tab_name, self) if self.app else None)
        self.btn_recover.pack(fill="x", padx=10, pady=5)
        
        self.btn_clear = ctk.CTkButton(self.frame_right, text="Clear Data", fg_color="#e74c3c", hover_color="#c0392b", command=self.clear_data)
        self.btn_clear.pack(fill="x", padx=10, pady=5)
        
        ctk.CTkFrame(self.frame_right, height=2, fg_color="gray").pack(fill="x", padx=10, pady=10)
        self.btn_save_tab = ctk.CTkButton(self.frame_right, text="Save Tab Configuration", fg_color="#27ae60", hover_color="#2ecc71", command=self.save_tab_config)
        self.btn_save_tab.pack(fill="x", padx=10, pady=(5, 5))
        self.btn_close_tab = ctk.CTkButton(self.frame_right, text="Close Tab", fg_color="#e74c3c", hover_color="#c0392b", command=lambda: self.app.close_tab(self))
        self.btn_close_tab.pack(fill="x", padx=10, pady=(5, 10))
        
        # --- MIDDLE PANEL: Views ---
        self.paned_window = tk.PanedWindow(self.frame_mid, orient=tk.VERTICAL, bg="#2b2b2b", sashwidth=6, sashrelief=tk.RAISED)
        self.paned_window.pack(fill="both", expand=True, padx=5, pady=5)
        
        # Plot Frame
        self.plot_frame = ctk.CTkFrame(self.paned_window, fg_color="transparent")
        
        # Dark mode styling for matplotlib
        plt.style.use('dark_background')
        self.fig, self.ax_mass = plt.subplots(figsize=(8, 6))
        self.fig.patch.set_facecolor('#2b2b2b')
        self.ax_mass.set_facecolor('#2b2b2b')
        self.ax_flow = self.ax_mass.twinx()
        
        self.ax_mass.set_title(f"{self.tab_name} - Live Weight Data (Not Recording)")
        self.ax_mass.set_xlabel("Time (minutes)")
        self.ax_mass.set_ylabel("Weight (g)", color='#3498db')
        self.ax_mass.tick_params(axis='y', labelcolor='#3498db')
        self.ax_mass.grid(True, alpha=0.3)
        
        self.ax_flow.set_ylabel("Flow Rate (g/min)", color='#e74c3c')
        self.ax_flow.tick_params(axis='y', labelcolor='#e74c3c')
        
        self.line_mass, = self.ax_mass.plot([], [], marker='o', linestyle='-', color='#3498db', label='Mass')
        self.line_flow, = self.ax_flow.plot([], [], marker='', linestyle='-', color='#e74c3c', label='Flow')
        self.line_fit, = self.ax_mass.plot([], [], marker='', linestyle='--', color='#f1c40f', label='Linear Fit', linewidth=2.5)
        self.ax_flow.set_visible(self.show_flow.get())
        self.fig.tight_layout()
        
        self.canvas = FigureCanvasTkAgg(self.fig, master=self.plot_frame)
        self.canvas.draw()
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        
        # Table Frame (Tkinter Treeview with custom dark style)
        self.table_frame = ctk.CTkFrame(self.paned_window, fg_color="transparent")
        
        style = tk.ttk.Style()
        style.theme_use("default")
        style.configure("Treeview", background="#2b2b2b", foreground="white", fieldbackground="#2b2b2b", borderwidth=0)
        style.map('Treeview', background=[('selected', '#1f538d')])
        style.configure("Treeview.Heading", background="#565b5e", foreground="white", relief="flat")
        style.map("Treeview.Heading", background=[('active', '#343638')])
        
        columns = ("Time", "Min", "Mass", "Flow")
        self.tree = tk.ttk.Treeview(self.table_frame, columns=columns, show="headings")
        
        self.tree.heading("Time", text="Timestamp")
        self.tree.heading("Min", text="Duration")
        self.tree.heading("Mass", text="Mass (g)")
        self.tree.heading("Flow", text="Flow (g/min)")
        
        self.tree.column("Time", width=120, anchor=tk.CENTER)
        self.tree.column("Min", width=60, anchor=tk.CENTER)
        self.tree.column("Mass", width=70, anchor=tk.CENTER)
        self.tree.column("Flow", width=80, anchor=tk.CENTER)
        
        scrollbar = tk.ttk.Scrollbar(self.table_frame, orient=tk.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscroll=scrollbar.set)
        scrollbar.pack(side="right", fill="y")
        self.tree.pack(side="left", fill="both", expand=True)
        
        self.paned_window.add(self.plot_frame, minsize=100, stretch="always")
        self.paned_window.add(self.table_frame, minsize=100, stretch="always")

    def save_tab_config(self):
        import datetime
        cfg = {
            "name": self.tab_name,
            "brand": self.combo_brand.get(),
            "port": self.combo_com.get(),
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        }
        self.app.config["saved_tabs"].append(cfg)
        self.app.save_config()
        messagebox.showinfo("Success", f"Tab Configuration Saved:\n\nName: {cfg['name']}\nBrand: {cfg['brand']}\nPort: {cfg['port']}")



    # --- Actions ---
    def on_filter_change(self, choice):
        if choice in ["Mean", "Median"]:
            self.lbl_filter_param.configure(text="Window Size (samples):")
            if choice == "Mean": self.ent_filter_param.delete(0, tk.END); self.ent_filter_param.insert(0, "20")
        elif choice == "EMA":
            self.lbl_filter_param.configure(text="Alpha (0.01 - 1.0):")
            self.ent_filter_param.delete(0, tk.END); self.ent_filter_param.insert(0, "0.1")
        elif choice == "Butterworth":
            self.lbl_filter_param.configure(text="Cutoff Freq (Hz):")
            self.ent_filter_param.delete(0, tk.END); self.ent_filter_param.insert(0, "0.01")
        elif choice == "Adaptive":
            self.lbl_filter_param.configure(text="Window (Auto-calculated)")
            
        self.last_flow_calc_n = 0

    def log_calibration(self):
        try:
            rpm = float(self.ent_rpm.get())
        except:
            messagebox.showerror("Error", "Please enter a valid Pump RPM.")
            return
            
        if len(self.flow_rates) == 0 or len(self.times_sec) == 0:
            messagebox.showerror("Error", "No data to log. Please wait for graph to update.")
            return
            
        current_time = self.times_sec[-1]
        stable_flows = []
        
        for i in range(len(self.times_sec)-1, -1, -1):
            if self.flow_rates[i] is not None:
                if current_time - self.times_sec[i] <= 30.0:
                    stable_flows.append(self.flow_rates[i])
                else:
                    break
                    
        if not stable_flows:
            messagebox.showerror("Error", "No stable flow rate to log. Please wait for graph to update.")
            return
            
        if current_time - self.times_sec[0] < 30.0:
            res = messagebox.askyesno("Warning", "You have less than 30 seconds of recorded data. The calibration point may not be fully stabilized yet.\n\nContinue logging anyway?")
            if not res: return
            
        flow = sum(stable_flows) / len(stable_flows)
        
        cal_path = os.path.join(os.getcwd(), f"{self.tab_name.replace(' ', '_')}_Calibration.csv")
        is_new = not os.path.exists(cal_path)
        
        try:
            with open(cal_path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if is_new:
                    writer.writerow(["Timestamp", "Pump RPM", "Avg Flow Rate (g/min)", "Averaging Window (s)"])
                ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                writer.writerow([ts, rpm, round(flow, 4), "30.0"])
            messagebox.showinfo("Success", f"Logged Calibration Point:\n\nRPM: {rpm}\nAvg Flow: {round(flow,4)} g/min (30s window)\n\nSaved to: {os.path.basename(cal_path)}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to log calibration:\n{e}")

    def refresh_com_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.combo_com.configure(values=ports if ports else ["No Ports Found"])
        if ports and not self.combo_com.get():
            self.combo_com.set(ports[0])

    def connect_serial(self, auto=False):
        port = self.combo_com.get()
        if not port or port == "No Ports Found" or port == "":
            if not auto:
                messagebox.showerror("Error", "Please select a valid COM Port before connecting.")
            return
            
        if self.ser and self.ser.is_open:
            self.ser.close()
            
        self.active_balance_brand = self.combo_brand.get()
        baud = 9600
        
        try:
            self.ser = serial.Serial(
                port=port, baudrate=baud, timeout=0.1,
                bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE,
                xonxoff=False, rtscts=False, dsrdtr=False
            )
            self.lbl_status.configure(text=f"Status: Connected ({port})", text_color="#2ecc71")
            self.btn_connect.configure(text="Connected", fg_color="#2ecc71", hover_color="#27ae60")
            
            if self.app:
                self.app.save_connection(self.tab_name, self.active_balance_brand, port)
                
        except Exception as e:
            self.lbl_status.configure(text="Status: Connection Failed", text_color="#e74c3c")
            if not auto:
                messagebox.showerror("Error", f"Could not connect to {port}:\n{e}")

    def disconnect_serial(self):
        if self.ser and self.ser.is_open:
            self.ser.close()
        self.ser = None
        self.lbl_status.configure(text="Status: Disconnected", text_color="gray")
        self.btn_connect.configure(text="Connect", fg_color="#3498db", hover_color="#2980b9")
        print(f"[{self.tab_name}] Disconnected.")

    def toggle_record(self):
        if not self.recording:
            # Start Recording
            if len(self.times_sec) == 0:
                self.start_time = time.time()
            if hasattr(self, 'wall_clock_min'):
                self.wall_clock_min.clear()
            self.recording = True
            self.last_activity_time = time.time()
            self.ax_mass.set_title(f"{self.tab_name} - Live Weight Data (RECORDING)")
            self.btn_record.configure(text="Stop Recording", fg_color="#e74c3c", hover_color="#c0392b")
            print(f"[{self.tab_name}] Recording started.")
        else:
            # Stop Recording
            self.recording = False
            self.ax_mass.set_title(f"{self.tab_name} - Live Weight Data (PAUSED)")
            self.btn_record.configure(text="Start Recording", fg_color="#2ecc71", hover_color="#27ae60")
            print(f"[{self.tab_name}] Recording stopped.")
            
        self.canvas.draw_idle()
            
    def tare_balance(self):
        if self.ser and self.ser.is_open:
            print(f"[{self.tab_name}] Sending Tare command...")
            self.ser.write(b"T\r\n")

    def clear_data(self):
        self.times_sec.clear()
        self.times_min.clear()
        self.weights.clear()
        self.timestamps.clear()
        self.flow_rates.clear()
        if hasattr(self, 'wall_clock_min'): self.wall_clock_min.clear()
        self.start_time = time.time()
        self.last_activity_time = time.time()
        self.last_recorded_time = 0.0
        self.last_flow_calc_n = 0
        
        self.line_mass.set_data([], [])
        self.line_flow.set_data([], [])
        self.line_fit.set_data([], [])
        self.canvas.draw_idle()
        
        for item in self.tree.get_children():
            self.tree.delete(item)
        print(f"[{self.tab_name}] Data cleared.")

    def set_float_entry(self, entry, val):
        entry.delete(0, tk.END)
        if val is not None:
            entry.insert(0, str(round(val, 2)))

    def apply_axis_limits(self):
        if not self.do_auto_scale_x.get():
            try:
                xmin = float(self.ent_xmin.get())
                xmax = float(self.ent_xmax.get())
                self.ax_mass.set_xlim(xmin, xmax)
            except Exception: pass
            
        if not self.do_auto_scale_y_mass.get():
            try:
                ymin = float(self.ent_ymin.get())
                ymax = float(self.ent_ymax.get())
                self.ax_mass.set_ylim(ymin, ymax)
            except Exception: pass
            
        if not self.do_auto_scale_y_flow.get():
            try:
                fmin = float(self.ent_flowmin.get())
                fmax = float(self.ent_flowmax.get())
                self.ax_flow.set_ylim(fmin, fmax)
            except Exception: pass
            
        self.canvas.draw_idle()

    def apply_linear_fit(self):
        if len(self.times_min) < 2:
            messagebox.showinfo("Fit", "Not enough data recorded yet.", parent=self.root)
            return
            
        try:
            start_min = float(self.ent_fit_start.get())
            end_min = float(self.ent_fit_end.get())
        except ValueError:
            messagebox.showerror("Error", "Invalid Fit Range. Please enter numbers.", parent=self.root)
            return
            
        valid_indices = [i for i, t in enumerate(self.times_min) if start_min <= t <= end_min]
        if len(valid_indices) < 2:
            messagebox.showinfo("Fit", "Not enough data points in the specified range.", parent=self.root)
            return
            
        x_vals = [self.times_min[i] for i in valid_indices]
        y_vals = [self.weights[i] for i in valid_indices]
        
        try:
            slope, intercept, r_value, p_value, std_err = scipy.stats.linregress(x_vals, y_vals)
            
            # Plot the line
            fit_y = [slope * x + intercept for x in x_vals]
            self.line_fit.set_data(x_vals, fit_y)
            self.canvas.draw_idle()
            
            # Append to notes
            note = f"\n[Linear Fit: {start_min} to {end_min} min] Avg Flow Rate: {slope:.4f} g/min | R²: {r_value**2:.4f}"
            self.txt_notes.insert(tk.END, note)
            self.unsaved_changes = True
            if self.app: self.app.set_unsaved_state(self.tab_name, True)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate fit:\n{e}", parent=self.root)

    def save_excel(self):
        if not self.times_sec:
            print(f"[{self.tab_name}] No data to save.")
            return
            
        dt_str = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
        filepath = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"{self.tab_name.replace(' ', '_')}_data_{dt_str}.xlsx",
            title=f"Save {self.tab_name} Excel Data",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if not filepath:
            return
            
        try:
            wb = openpyxl.Workbook()
            
            # Documentation Sheet
            doc_ws = wb.active
            doc_ws.title = "Documentation"
            
            doc_ws["A1"] = f"{self.tab_name} - Experiment Documentation"
            doc_ws["A1"].font = Font(size=14, bold=True)
            
            doc_ws["A3"] = "Export Date:"
            doc_ws["B3"] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            doc_ws["A4"] = "Balance Brand:"
            doc_ws["B4"] = self.active_balance_brand
            
            doc_ws["A5"] = "COM Port:"
            doc_ws["B5"] = self.combo_com.get()
            
            doc_ws["A6"] = "Recording Interval (s):"
            doc_ws["B6"] = self.ent_interval.get()
            
            doc_ws["A7"] = "Derivative SavGol Window:"
            doc_ws["B7"] = self.ent_savgol_win.get()
            
            doc_ws["A8"] = "Derivative SavGol Poly:"
            doc_ws["B8"] = self.ent_savgol_poly.get()
            
            doc_ws["A9"] = "Secondary Filter Type:"
            doc_ws["B9"] = self.combo_filter.get()
            
            doc_ws["A10"] = "Secondary Filter Window:"
            doc_ws["B10"] = self.ent_filter_param.get()
            
            doc_ws["A12"] = "Experiment Notes:"
            doc_ws["A12"].font = Font(bold=True)
            
            notes = self.txt_notes.get("0.0", tk.END).strip()
            doc_ws["A13"] = notes
            doc_ws["A13"].alignment = Alignment(wrap_text=True, vertical="top")
            
            doc_ws.column_dimensions["A"].width = 25
            doc_ws.column_dimensions["B"].width = 30
            doc_ws.row_dimensions[13].height = 80
            
            # Data Sheet
            data_ws = wb.create_sheet("Data")
            headers = ["Timestamp", "Duration (min)", "Mass (g)", "Flow Rate (g/min)"]
            data_ws.append(headers)
            for cell in data_ws[1]:
                cell.font = Font(bold=True)
                
            for i in range(len(self.times_sec)):
                ts = self.timestamps[i]
                dur = round(self.times_min[i], 5)
                mass = self.weights[i]
                flw = self.flow_rates[i] if i < len(self.flow_rates) and self.flow_rates[i] is not None else ""
                data_ws.append([ts, dur, mass, flw])
                
            # Chart Sheet
            chart_ws = wb.create_sheet("Chart")
            chart = ScatterChart()
            chart.title = f"{self.tab_name} - Weight vs. Time"
            chart.x_axis.title = "Duration (min)"
            chart.y_axis.title = "Mass (g)"
            chart.style = 13
            
            max_row = len(self.times_sec) + 1
            if max_row > 1:
                xvalues = Reference(data_ws, min_col=2, min_row=2, max_row=max_row)
                yvalues = Reference(data_ws, min_col=3, min_row=2, max_row=max_row)
                series = Series(yvalues, xvalues, title="Recorded Weight")
                chart.series.append(series)
                
            chart_ws.add_chart(chart, "A1")
                
            wb.save(filepath)
            
            # Reset unsaved state
            self.unsaved_changes = False
            if self.app:
                self.app.set_unsaved_state(self.tab_name, False)
                
            print(f"[{self.tab_name}] Data saved to {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save Excel:\n{e}")

    def scan_backup_sessions(self):
        if not os.path.exists(self.backup_path):
            return []
            
        sessions = []
        try:
            with open(self.backup_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                rows = list(reader)
                
            current_session = None
            last_dur = -1.0
            
            for idx, row in enumerate(rows):
                if len(row) >= 4:
                    try:
                        ts = row[0]
                        dur = float(row[1])
                        
                        if current_session is None or dur < last_dur:
                            if current_session:
                                current_session["end_idx"] = idx - 1
                                sessions.append(current_session)
                                
                            current_session = {
                                "start_idx": idx,
                                "start_ts": ts,
                                "end_ts": ts,
                                "points": 0
                            }
                            
                        current_session["end_ts"] = ts
                        current_session["points"] += 1
                        last_dur = dur
                    except: pass
                    
            if current_session:
                current_session["end_idx"] = len(rows) - 1
                sessions.append(current_session)
                
        except Exception as e:
            print(f"[{self.tab_name}] Error scanning backup: {e}")
            
        return sessions

    def get_wall_clock_data(self):
        import datetime
        if not hasattr(self, 'wall_clock_min'):
            self.wall_clock_min = []
            
        if len(self.wall_clock_min) < len(self.timestamps):
            try:
                if len(self.wall_clock_min) == 0:
                    self._t0 = datetime.datetime.strptime(self.timestamps[0], "%Y-%m-%d %H:%M:%S")
                    
                for i in range(len(self.wall_clock_min), len(self.timestamps)):
                    t = datetime.datetime.strptime(self.timestamps[i], "%Y-%m-%d %H:%M:%S")
                    self.wall_clock_min.append((t - self._t0).total_seconds() / 60.0)
            except Exception as e:
                while len(self.wall_clock_min) < len(self.timestamps):
                    self.wall_clock_min.append(self.times_min[len(self.wall_clock_min)])
                    
        return self.wall_clock_min

    def restore_from_backup(self, session_data=None):
        if not os.path.exists(self.backup_path):
            print(f"[{self.tab_name}] No backup file found to restore.")
            return
            
        try:
            with open(self.backup_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                rows = list(reader)
                
            start_idx = 0
            end_idx = len(rows) - 1
            if session_data:
                start_idx = session_data["start_idx"]
                end_idx = session_data["end_idx"]
                
            for idx in range(start_idx, end_idx + 1):
                row = rows[idx]
                if len(row) >= 4:
                    try:
                        ts = row[0]
                        dur = float(row[1])
                        mass = float(row[2])
                        flw_str = row[3]
                        flw = float(flw_str) if flw_str else None
                        
                        self.timestamps.append(ts)
                        curr = dur * 60.0
                        self.times_sec.append(curr)
                        self.times_min.append(dur)
                        self.weights.append(mass)
                        if flw is not None:
                            self.flow_rates.append(flw)
                    except: pass
                        
            if self.times_sec:
                self.start_time = time.time() - self.times_sec[-1]
                self.last_recorded_time = self.times_sec[-1]
                self.unsaved_changes = True
                print(f"[{self.tab_name}] Restored {len(self.times_sec)} points from backup.")
                
        except Exception as e:
            print(f"[{self.tab_name}] Error restoring backup: {e}")

    def save_graph(self):
        dt_str = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
        filepath = filedialog.asksaveasfilename(
            defaultextension=".png",
            initialfile=f"{self.tab_name.replace(' ', '_')}_graph_{dt_str}.png",
            title=f"Save {self.tab_name} Graph",
            filetypes=[("PNG Image", "*.png"), ("PDF Document", "*.pdf")]
        )
        if filepath:
            try:
                self.fig.savefig(filepath)
                print(f"[{self.tab_name}] Graph saved to {filepath}")
                messagebox.showinfo("Success", f"Graph successfully saved to:\n{filepath}")
            except Exception as e:
                print(f"[{self.tab_name}] Error saving graph: {e}")
                messagebox.showerror("Error", f"Failed to save graph:\n{e}")

    # --- Serial Logic ---
    def serial_worker(self):
        """ Runs constantly in background thread """
        while self.is_running_thread:
            if not self.ser or not self.ser.is_open:
                time.sleep(0.5)
                continue
                
            try:
                # Poll MT Balance if needed
                if self.active_balance_brand == "Mettler Toledo":
                    if self.ser.in_waiting == 0:
                        self.ser.write(b"SI\r\n")
                elif self.active_balance_brand == "Ohaus Adventurer":
                    if self.ser.in_waiting == 0:
                        self.ser.write(b"IP\r\n")
                        
                response = self.ser.readline()
                if response:
                    decoded = response.decode('ascii', errors='ignore').strip()
                    if not decoded:
                        continue
                        
                    if self.active_balance_brand == "Mettler Toledo":
                        status, value, unit = parse_mt_sics_response(decoded)
                    elif self.active_balance_brand == "Ohaus Adventurer":
                        status, value, unit = parse_ohaus_response(decoded)
                    else:
                        status, value, unit = parse_bonvoisin_response(decoded)
                    
                    if value is not None and self.recording:
                        # Grab UI variables safely
                        try:
                            intrvl = float(self.ent_interval.get())
                            self.sampling_interval = max(0.0, intrvl)
                        except:
                            self.sampling_interval = 1.0
                        
                        curr = time.time() - self.start_time
                        if (curr - self.last_recorded_time >= self.sampling_interval) or len(self.times_sec) == 0:
                            self.last_recorded_time = curr
                            
                            self.times_sec.append(curr)
                            self.times_min.append(curr / 60.0)
                            self.weights.append(value)
                            
                            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
                            self.timestamps.append(ts)
                            
                            # Backup to CSV
                            try:
                                is_new = not os.path.exists(self.backup_path)
                                with open(self.backup_path, 'a', newline='', encoding='utf-8') as f:
                                    writer = csv.writer(f)
                                    if is_new:
                                        writer.writerow(["Timestamp", "Duration (min)", "Mass (g)", "Flow Rate (g/min)"])
                                    writer.writerow([ts, round(curr / 60.0, 5), value, ""])
                            except Exception as e:
                                print(f"[{self.tab_name}] Backup error: {e}")
                                
                            if not self.unsaved_changes:
                                self.unsaved_changes = True
                                if self.app:
                                    self.app.set_unsaved_state(self.tab_name, True)
                            
                            if not self.unit_set:
                                self.current_unit = unit
                                self.unit_set = True
                                
            except Exception as e:
                time.sleep(0.1)

    # --- GUI Update Loop ---
    def calculate_flows(self):
        n = len(self.times_sec)
        while len(self.flow_rates) < n:
            self.flow_rates.append(None)
            
        valid_indices = []
        raw_flows = [None] * n
        
        if n >= 5:
            try:
                sg_win = max(5, int(self.ent_savgol_win.get()))
                if sg_win % 2 == 0: sg_win += 1
                sg_poly = min(sg_win - 1, int(self.ent_savgol_poly.get()))
                
                import numpy as np
                import scipy.signal
                
                if n >= sg_win:
                    w_sm = scipy.signal.savgol_filter(self.weights, window_length=sg_win, polyorder=sg_poly)
                    deriv_raw = np.gradient(w_sm, self.times_sec)
                else:
                    deriv_raw = np.gradient(self.weights, self.times_sec)
                    
                for i in range(n):
                    f = deriv_raw[i] * 60.0
                    raw_flows[i] = f
                    valid_indices.append(i)
            except Exception as e:
                import numpy as np
                deriv_raw = np.gradient(self.weights, self.times_sec)
                for i in range(n):
                    f = deriv_raw[i] * 60.0
                    raw_flows[i] = f
                    valid_indices.append(i)
        elif n > 1:
            import numpy as np
            deriv_raw = np.gradient(self.weights, self.times_sec)
            for i in range(n):
                f = deriv_raw[i] * 60.0
                raw_flows[i] = f
                valid_indices.append(i)
                
        if not valid_indices:
            return
            
        filter_type = self.combo_filter.get()
        
        if filter_type in ['Mean', 'Median']:
            try: win = max(1, int(self.ent_filter_param.get()))
            except: win = 20
            
            for v_idx in range(len(valid_indices)):
                if v_idx >= win - 1:
                    i = valid_indices[v_idx]
                    window_indices = valid_indices[v_idx - win + 1 : v_idx + 1]
                    vals = [raw_flows[idx] for idx in window_indices]
                    if filter_type == 'Mean':
                        self.flow_rates[i] = statistics.mean(vals)
                    else:
                        self.flow_rates[i] = statistics.median(vals)
                        
        elif filter_type == 'EMA':
            try: alpha = float(self.ent_filter_param.get())
            except: alpha = 0.1
            alpha = max(0.01, min(1.0, alpha))
            
            self.flow_rates[valid_indices[0]] = raw_flows[valid_indices[0]]
            for v_idx in range(1, len(valid_indices)):
                i = valid_indices[v_idx]
                prev_i = valid_indices[v_idx-1]
                if self.flow_rates[prev_i] is not None:
                    self.flow_rates[i] = alpha * raw_flows[i] + (1 - alpha) * self.flow_rates[prev_i]
                else:
                    self.flow_rates[i] = raw_flows[i]
                    
        elif filter_type == 'Butterworth':
            try: cutoff = float(self.ent_filter_param.get())
            except: cutoff = 0.05
            
            try:
                # Approximate sampling frequency
                fs = 1.0 / max(0.1, getattr(self, 'sampling_interval', 1.0))
                nyq = 0.5 * fs
                normal_cutoff = cutoff / nyq
                
                if len(valid_indices) > 15 and normal_cutoff < 1.0:
                    import scipy.signal
                    b, a = scipy.signal.butter(2, normal_cutoff, btype='low', analog=False)
                    raw_vals = [raw_flows[idx] for idx in valid_indices]
                    filtered_vals = scipy.signal.filtfilt(b, a, raw_vals)
                    for v_idx, val in enumerate(filtered_vals):
                        self.flow_rates[valid_indices[v_idx]] = val
                else:
                    for i in valid_indices:
                        self.flow_rates[i] = raw_flows[i]
            except Exception as e:
                print(f"Butterworth error: {e}")
                for i in valid_indices:
                    self.flow_rates[i] = raw_flows[i]
                    
        elif filter_type == 'Adaptive':
            try: rpm = float(self.ent_rpm.get())
            except: rpm = 30.0
            if rpm <= 0: rpm = 1.0
            
            try: rollers = float(self.ent_rollers.get())
            except: rollers = 3.0
            if rollers <= 0: rollers = 1.0
            
            pulse_hz = (rpm / 60.0) * rollers
            period_sec = 1.0 / pulse_hz
            win = max(1, int(round(period_sec / max(0.1, getattr(self, 'sampling_interval', 1.0)))))
            
            for v_idx in range(len(valid_indices)):
                if v_idx >= win - 1:
                    i = valid_indices[v_idx]
                    window_indices = valid_indices[v_idx - win + 1 : v_idx + 1]
                    vals = [raw_flows[idx] for idx in window_indices]
                    self.flow_rates[i] = statistics.mean(vals)

    def update_gui(self):
        if not self.is_running_thread:
            return
            
        n = len(self.times_sec)
        n_items = len(self.tree.get_children())
        
        needs_redraw = False
        if n > getattr(self, 'last_flow_calc_n', 0):
            self.calculate_flows()
            self.last_flow_calc_n = n
            needs_redraw = True
            
            # --- Auto-Stop Logic ---
            if self.recording and getattr(self, 'do_auto_stop', None) and self.do_auto_stop.get():
                if len(self.flow_rates) > 0 and self.flow_rates[-1] is not None:
                    current_flow = self.flow_rates[-1]
                    try: thresh = float(self.ent_auto_stop_thresh.get())
                    except: thresh = 0.1
                    
                    if abs(current_flow) > thresh:
                        self.last_activity_time = time.time()
                    else:
                        try: stop_min = float(self.ent_auto_stop_min.get())
                        except: stop_min = 5.0
                        
                        if time.time() - getattr(self, 'last_activity_time', time.time()) > (stop_min * 60.0):
                            print(f"[{self.tab_name}] Auto-stop triggered!")
                            self.toggle_record()
                            self.txt_notes.insert(tk.END, f"\n[Auto-Stop] Triggered at {datetime.datetime.now().strftime('%H:%M:%S')} due to flow rate < {thresh} for {stop_min} min.")
            # -----------------------
            
        if needs_redraw or n > n_items:
            try:
                # --- UPDATE PLOT ---
                f_v = []
                if len(self.times_min) > 0:
                    if getattr(self, 'var_xaxis', None) and self.var_xaxis.get() == "Timestamp":
                        x_data = self.get_wall_clock_data()
                        self.ax_mass.set_xlabel("Elapsed Timestamp (min)")
                    else:
                        x_data = self.times_min
                        self.ax_mass.set_xlabel("Duration (min)")
                        
                    self.line_mass.set_data(list(x_data), list(self.weights))
                    
                    f_v = [f for f in self.flow_rates if f is not None]
                    if self.show_flow.get() and len(f_v) > 0:
                        t_v = [x_data[i] for i in range(len(x_data)) if self.flow_rates[i] is not None]
                        self.line_flow.set_data(list(t_v), list(f_v))
                
                if self.unit_set:
                    self.ax_mass.set_ylabel(f"Weight ({self.current_unit})")
                    
                if self.do_auto_scale_x.get() or self.do_auto_scale_y_mass.get() or self.do_auto_scale_y_flow.get():
                    self.ax_mass.relim()
                    if self.do_auto_scale_x.get():
                        self.ax_mass.autoscale(enable=True, axis='x', tight=None)
                    if self.do_auto_scale_y_mass.get():
                        self.ax_mass.autoscale(enable=True, axis='y', tight=None)
                        
                    if self.show_flow.get() and len(f_v) > 0:
                        self.ax_flow.relim()
                        if self.do_auto_scale_y_flow.get():
                            self.ax_flow.autoscale(enable=True, axis='y', tight=None)
                        
                    xl = self.ax_mass.get_xlim()
                    yl = self.ax_mass.get_ylim()
                    if self.do_auto_scale_x.get():
                        self.set_float_entry(self.ent_xmin, xl[0])
                        self.set_float_entry(self.ent_xmax, xl[1])
                    if self.do_auto_scale_y_mass.get():
                        self.set_float_entry(self.ent_ymin, yl[0])
                        self.set_float_entry(self.ent_ymax, yl[1])
                    
                    if self.show_flow.get() and len(f_v) > 0:
                        fl = self.ax_flow.get_ylim()
                        if self.do_auto_scale_y_flow.get():
                            self.set_float_entry(self.ent_flowmin, fl[0])
                            self.set_float_entry(self.ent_flowmax, fl[1])
                        
                self.canvas.draw_idle()
                
                # --- UPDATE TABLE ---
                start_idx = n_items
                insert_limit = 500
                end_idx = min(n, start_idx + insert_limit)
                
                for i in range(start_idx, end_idx):
                    if i < len(self.flow_rates) and self.flow_rates[i] is not None:
                        flow_str = f"{self.flow_rates[i]:.4f}"
                    else: 
                        flow_str = ""
                        
                    values = (
                        self.timestamps[i].split(" ")[1],
                        f"{self.times_min[i]:.3f}",
                        f"{self.weights[i]:.3f}",
                        flow_str
                    )
                    
                    self.tree.insert("", tk.END, values=values)
                        
                if len(self.tree.get_children()) > 0:
                    self.tree.yview_moveto(1)
            except Exception as e:
                print(f"[{self.tab_name}] Error updating GUI: {e}")

    def close(self):
        self.is_running_thread = False
        if self.ser and self.ser.is_open:
            self.ser.close()


class MultiBalanceApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Multi-Balance Data Streamer")
        self.root.geometry("1400x800")
        ctk.set_appearance_mode("Dark")
        ctk.set_default_color_theme("blue")
        
        self.tabs = []
        
        # Load Config
        import json
        self.config_path = os.path.join(os.getcwd(), "config.json")
        self.config = {"auto_connect": True, "ui_scale": 1.0, "known_balances": {}, "saved_tabs": []}
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, "r") as f:
                    self.config.update(json.load(f))
            except: pass
            
        # Apply Scaling
        try:
            scale = float(self.config.get("ui_scale", 1.0))
            ctk.set_widget_scaling(scale)
            ctk.set_window_scaling(scale)
        except Exception as e: print(e)
        
        # --- Top Control Bar ---
        self.top_frame = ctk.CTkFrame(self.root)
        self.top_frame.pack(side="top", fill="x", padx=10, pady=10)
        
        ctk.CTkLabel(self.top_frame, text="New Tab Name:", font=ctk.CTkFont(weight="bold")).pack(side="left", padx=5)
        self.ent_tab_name = ctk.CTkEntry(self.top_frame, width=150)
        self.ent_tab_name.insert(0, "Balance 1")
        self.ent_tab_name.pack(side="left", padx=5)
        
        self.btn_add_tab = ctk.CTkButton(self.top_frame, text="Add Tab", width=80, command=self.add_tab)
        self.btn_add_tab.pack(side="left", padx=5)
        
        self.btn_load_tab = ctk.CTkButton(self.top_frame, text="📂 Load Saved Tab", width=120, command=self.open_tab_manager)
        self.btn_load_tab.pack(side="left", padx=5)
        
        self.btn_rename_tab = ctk.CTkButton(self.top_frame, text="Rename Tab", fg_color="#f39c12", hover_color="#d35400", command=self.rename_tab)
        self.btn_rename_tab.pack(side="left", padx=5)
        
        self.btn_delete_tab = ctk.CTkButton(self.top_frame, text="Delete Tab", fg_color="#e74c3c", hover_color="#c0392b", command=self.delete_tab)
        self.btn_delete_tab.pack(side="left", padx=5)
        
        self.btn_settings = ctk.CTkButton(self.top_frame, text="⚙ Settings", width=80, fg_color="transparent", border_width=1, text_color=("gray10", "gray90"), command=self.open_settings)
        self.btn_settings.pack(side="right", padx=5)
        
        # --- Notebook (Tabview) ---
        self.tabview = ctk.CTkTabview(self.root)
        self.tabview.pack(side="top", fill="both", expand=True, padx=10, pady=(0, 10))
        
        # Auto Connect from Memory
        has_known = False
        if self.config.get("auto_connect", True):
            for name, details in self.config["known_balances"].items():
                tab = self.add_tab(name=name, brand=details["brand"], port=details["port"])
                
                # Allow UI to build before connecting
                self.root.after(500, lambda t=tab: t.connect_serial(auto=True))
                has_known = True
                
        if not has_known:
            self.add_tab()
            
        # Start Master GUI Loop
        self.root.after(100, self.master_gui_loop)

    def save_config(self):
        import json
        try:
            with open(self.config_path, "w") as f:
                json.dump(self.config, f, indent=4)
        except: pass

    def save_connection(self, name, brand, port):
        if "known_balances" not in self.config:
            self.config["known_balances"] = {}
        self.config["known_balances"][name] = {"brand": brand, "port": port, "unsaved": False}
        self.save_config()

    def set_unsaved_state(self, name, state):
        if "known_balances" in self.config and name in self.config["known_balances"]:
            self.config["known_balances"][name]["unsaved"] = state
            self.save_config()

    def add_tab(self, name=None, brand=None, port=None):
        if not name:
            name = self.ent_tab_name.get()
            if not name:
                name = f"Balance {len(self.tabs) + 1}"
            
        self.tabview.add(name)
        new_tab = BalanceTab(self.tabview.tab(name), self.root, name, self)
        new_tab.pack(fill="both", expand=True)
        
        if brand:
            new_tab.combo_brand.set(brand)
        if port:
            new_tab.combo_com.set(port)
            
        self.tabs.append(new_tab)
        
        self.ent_tab_name.delete(0, tk.END)
        self.ent_tab_name.insert(0, f"Balance {len(self.tabs) + 1}")
        
        return new_tab

    def rename_tab(self):
        current_name = self.tabview.get()
        if not current_name:
            return
            
        dialog = ctk.CTkInputDialog(text=f"Enter new name for '{current_name}':", title="Rename Tab")
        new_name = dialog.get_input()
        
        if not new_name or new_name == current_name:
            return
            
        if new_name in [t.tab_name for t in self.tabs]:
            messagebox.showerror("Error", f"A tab named '{new_name}' already exists.")
            return
            
        try:
            self.tabview.rename(current_name, new_name)
            self.tabview._current_name = new_name
            self.tabview.set(new_name)
        except Exception as e:
            messagebox.showerror("Error", f"Failed to rename tab: {e}")
            return
            
        tab = next((t for t in self.tabs if t.tab_name == current_name), None)
        if tab:
            tab.tab_name = new_name
            tab.ax_mass.set_title(f"{new_name} - Live Weight Data")
            tab.canvas.draw_idle()
            
            old_backup = tab.backup_path
            backup_dir = os.path.join(os.getcwd(), "Data Backups")
            os.makedirs(backup_dir, exist_ok=True)
            tab.backup_path = os.path.join(backup_dir, f"{new_name.replace(' ', '_')}_Backup.csv")
            if os.path.exists(old_backup):
                try: os.rename(old_backup, tab.backup_path)
                except: pass
                
        if "known_balances" in self.config and current_name in self.config["known_balances"]:
            self.config["known_balances"][new_name] = self.config["known_balances"].pop(current_name)
            self.save_config()

    def delete_tab(self):
        current_name = self.tabview.get()
        if not current_name:
            return
            
        confirm = messagebox.askyesno("Confirm Delete", f"Are you sure you want to delete the tab '{current_name}'?\nThis will disconnect the balance and remove it from memory.", parent=self.root)
        if not confirm:
            return
            
        tab = next((t for t in self.tabs if t.tab_name == current_name), None)
        if tab:
            if hasattr(tab, 'save_tab_settings'):
                tab.save_tab_settings()
            tab.close()
            self.tabs.remove(tab)
            
        try:
            self.tabview.delete(current_name)
        except Exception as e:
            print(f"Error deleting tab from view: {e}")
            
        if "known_balances" in self.config and current_name in self.config["known_balances"]:
            del self.config["known_balances"][current_name]
            self.save_config()

    def prompt_session_recovery(self, name, tab):
        sessions = tab.scan_backup_sessions()
        if not sessions:
            self.set_unsaved_state(name, False)
            messagebox.showinfo("Recovery", f"No recovery sessions found for '{name}'.", parent=self.root)
            return
            
        if len(sessions) == 1:
            do_restore = messagebox.askyesno("Unsaved Data", f"You have an unsaved session for '{name}'.\n\nWould you like to restore it?", parent=self.root)
            if do_restore:
                tab.restore_from_backup(sessions[0])
            else:
                self.set_unsaved_state(name, False)
            return
            
        rec_win = ctk.CTkToplevel(self.root)
        rec_win.title(f"Restore Session - {name}")
        rec_win.geometry("550x400")
        rec_win.transient(self.root)
        rec_win.grab_set()
        
        ctk.CTkLabel(rec_win, text=f"Unsaved Data Found: {name}", font=ctk.CTkFont(size=16, weight="bold")).pack(pady=10)
        ctk.CTkLabel(rec_win, text="Multiple recording sessions were found. Please select which one to restore:").pack(pady=5)
        
        scroll_frame = ctk.CTkScrollableFrame(rec_win, height=200)
        scroll_frame.pack(fill="both", expand=True, padx=20, pady=10)
        
        selected_session = tk.IntVar(value=len(sessions)-1)
        
        for i, s in enumerate(sessions):
            pts = s['points']
            txt = f"Session {i+1} | {s['start_ts']} -> {s['end_ts']} | {pts} pts"
            rb = ctk.CTkRadioButton(scroll_frame, text=txt, variable=selected_session, value=i)
            rb.pack(anchor="w", pady=5)
            
        def on_restore():
            idx = selected_session.get()
            if idx >= 0 and idx < len(sessions):
                tab.restore_from_backup(sessions[idx])
            else:
                self.set_unsaved_state(name, False)
            rec_win.destroy()
            
        def on_ignore():
            self.set_unsaved_state(name, False)
            rec_win.destroy()
            
        btn_frame = ctk.CTkFrame(rec_win, fg_color="transparent")
        btn_frame.pack(pady=10)
        
        ctk.CTkButton(btn_frame, text="Restore Selected", fg_color="#27ae60", hover_color="#219150", command=on_restore).pack(side="left", padx=10)
        ctk.CTkButton(btn_frame, text="Ignore All", fg_color="#e74c3c", hover_color="#c0392b", command=on_ignore).pack(side="right", padx=10)

    def open_settings(self):
        settings_win = ctk.CTkToplevel(self.root)
        settings_win.title("Settings")
        settings_win.geometry("450x400")
        settings_win.transient(self.root)
        settings_win.grab_set()

        ctk.CTkLabel(settings_win, text="Settings", font=ctk.CTkFont(size=18, weight="bold")).pack(pady=15)

        # UI Scale Setting
        ctk.CTkLabel(settings_win, text="UI Scale (Requires Restart):", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=20, pady=(0, 5))
        
        scale_var = ctk.StringVar(value=str(int(self.config.get("ui_scale", 1.0) * 100)) + "%")
        combo_scale = ctk.CTkComboBox(settings_win, values=["80%", "100%", "125%", "150%"], variable=scale_var, command=self._update_ui_scale)
        combo_scale.pack(anchor="w", padx=20, pady=(0, 10))

        # Auto connect toggle
        auto_var = tk.BooleanVar(value=self.config.get("auto_connect", True))
        chk_auto = ctk.CTkCheckBox(settings_win, text="Enable Auto-Connect on Startup", variable=auto_var, 
                                   command=lambda: self._update_auto_connect(auto_var.get()))
        chk_auto.pack(anchor="w", padx=20, pady=5)

        ctk.CTkFrame(settings_win, height=2, fg_color="gray").pack(fill="x", pady=10, padx=20)

        ctk.CTkLabel(settings_win, text="Known Balances Memory:", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", padx=20)

        frame_mem = ctk.CTkFrame(settings_win, fg_color="transparent")
        frame_mem.pack(fill="both", expand=True, padx=20, pady=5)

        for name, details in self.config.get("known_balances", {}).items():
            f = ctk.CTkFrame(frame_mem)
            f.pack(fill="x", pady=2)
            ctk.CTkLabel(f, text=f"{name} ({details['brand']} on {details['port']})").pack(side="left", padx=10)
            btn_del = ctk.CTkButton(f, text="Forget", width=60, fg_color="#e74c3c", hover_color="#c0392b", command=lambda n=name, w=f: self._forget_balance(n, w))
            btn_del.pack(side="right", padx=10, pady=5)

        ctk.CTkButton(settings_win, text="Close", command=settings_win.destroy).pack(pady=15)

    def _update_ui_scale(self, val):
        try:
            scale_float = float(val.replace("%", "")) / 100.0
            self.config["ui_scale"] = scale_float
            self.save_config()
            messagebox.showinfo("Restart Required", "UI Scale setting saved.\n\nPlease restart the application for the changes to take effect.")
        except: pass

    def _update_auto_connect(self, val):
        self.config["auto_connect"] = val
        self.save_config()

    def _forget_balance(self, name, widget):
        if name in self.config["known_balances"]:
            del self.config["known_balances"][name]
            self.save_config()
            widget.destroy()
        
    def master_gui_loop(self):
        for tab in self.tabs:
            if tab.is_running_thread:
                tab.update_gui()
        self.root.after(100, self.master_gui_loop)
        
    def open_tab_manager(self):
        manager_win = ctk.CTkToplevel(self.root)
        manager_win.title("Tab Configuration Manager")
        manager_win.geometry("500x350")
        manager_win.grab_set()
        
        manager_win.grid_columnconfigure(0, weight=1)
        manager_win.grid_columnconfigure(1, weight=1)
        manager_win.grid_rowconfigure(0, weight=1)
        
        list_frame = ctk.CTkScrollableFrame(manager_win)
        list_frame.grid(row=0, column=0, sticky="nsew", padx=10, pady=10)
        
        preview_frame = ctk.CTkFrame(manager_win)
        preview_frame.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        
        lbl_preview = ctk.CTkLabel(preview_frame, text="Select a tab to preview", justify="left")
        lbl_preview.pack(pady=10, padx=10)
        
        btn_load = ctk.CTkButton(preview_frame, text="Load Tab", state="disabled")
        btn_load.pack(fill="x", pady=5, padx=10)
        
        btn_delete = ctk.CTkButton(preview_frame, text="Delete", fg_color="#e74c3c", hover_color="#c0392b", state="disabled")
        btn_delete.pack(fill="x", pady=5, padx=10)
        
        selected_idx = [None]
        
        def on_select(idx):
            selected_idx[0] = idx
            tab_data = self.config["saved_tabs"][idx]
            lbl_preview.configure(text=f"Name: {tab_data['name']}\nBrand: {tab_data['brand']}\nPort: {tab_data['port']}\nSaved:\n{tab_data['timestamp']}")
            btn_load.configure(state="normal")
            btn_delete.configure(state="normal")
            
        def build_list():
            for widget in list_frame.winfo_children():
                widget.destroy()
            for i, tab_data in enumerate(self.config["saved_tabs"]):
                btn = ctk.CTkButton(list_frame, text=f"{tab_data['name']}\n({tab_data['timestamp']})", 
                                    command=lambda idx=i: on_select(idx), anchor="w")
                btn.pack(fill="x", pady=2)
            lbl_preview.configure(text="Select a tab to preview")
            btn_load.configure(state="disabled")
            btn_delete.configure(state="disabled")
            selected_idx[0] = None
            
        def do_load():
            idx = selected_idx[0]
            if idx is not None:
                tab_data = self.config["saved_tabs"][idx]
                self.add_tab_with_settings(tab_data['name'], tab_data['brand'], tab_data['port'])
                manager_win.destroy()
                
        def do_delete():
            idx = selected_idx[0]
            if idx is not None:
                del self.config["saved_tabs"][idx]
                self.save_config()
                build_list()
                
        btn_load.configure(command=do_load)
        btn_delete.configure(command=do_delete)
        build_list()

    def add_tab_with_settings(self, name, brand, port):
        self.tabview.add(name)
        tab_frame = self.tabview.tab(name)
        tab = BalanceTab(tab_frame, self, name)
        
        tab.combo_brand.set(brand)
        if port not in tab.combo_com.cget("values"):
            vals = list(tab.combo_com.cget("values"))
            if "No Ports Found" in vals: vals.remove("No Ports Found")
            vals.append(port)
            tab.combo_com.configure(values=vals)
        tab.combo_com.set(port)
        
        self.tabs.append(tab)
        self.tabview.set(name)
        
    def close_tab(self, tab_obj):
        if tab_obj.recording:
            messagebox.showwarning("Warning", "Stop recording before closing the tab.")
            return
        if tab_obj.ser and tab_obj.ser.is_open:
            tab_obj.disconnect_serial()
        tab_obj.is_running_thread = False
        
        name = tab_obj.tab_name
        self.tabs.remove(tab_obj)
        self.tabview.delete(name)

    def on_closing(self):
        for tab in self.tabs:
            if hasattr(tab, 'save_tab_settings'):
                tab.save_tab_settings()
            tab.close()
        self.save_config()
        self.root.destroy()
        print("Application closed.")

if __name__ == "__main__":
    import ctypes
    try:
        myappid = 'candr.balance_streamer.1.0'
        ctypes.windll.shell32.SetCurrentProcessExplicitAppUserModelID(myappid)
    except Exception: pass
    
    root = ctk.CTk()
    app = MultiBalanceApp(root)
    root.protocol("WM_DELETE_WINDOW", app.on_closing)
    root.mainloop()
