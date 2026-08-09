import sys
import os

# Ensure the working directory is the location of the script
script_dir = os.path.dirname(os.path.abspath(__file__))
if os.getcwd() != script_dir:
    os.chdir(script_dir)

import serial
import serial.tools.list_ports
import time
import re
import csv
import datetime
import threading
import statistics
import json

from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, 
                             QHBoxLayout, QLabel, QLineEdit, QPushButton, 
                             QComboBox, QCheckBox, QTextEdit, QGroupBox, 
                             QMessageBox, QFileDialog, QTabWidget, QDialog, 
                             QGridLayout, QSplitter, QScrollArea, QTableWidget, 
                             QTableWidgetItem, QAbstractItemView, QHeaderView, QListWidget,
                             QButtonGroup, QRadioButton, QFormLayout)
from PyQt6.QtCore import Qt, QTimer, pyqtSignal, QObject, QThread
from PyQt6.QtGui import QFont, QColor

import qdarktheme

import matplotlib
matplotlib.use('QtAgg')
import matplotlib.pyplot as plt
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg

import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.chart import ScatterChart, Reference, Series
import scipy.stats
import scipy.signal

# --- Global Parsers ---
def parse_bonvoisin_response(response_str):
    match = re.search(r"([-+]?(?:\d+\.\d+|\d+))", response_str)
    if match:
        value = float(match.group(1))
        unit_match = re.search(r"([a-zA-Z]+)", response_str)
        unit = unit_match.group(1) if unit_match else "g"
        return "S", value, unit
    return None, None, None

mt_response_pattern = re.compile(r"S\s+([SD])\s+([ \-0-9\.]+)\s+([a-zA-Z]+)")
def parse_mt_sics_response(response_str):
    match = mt_response_pattern.search(response_str)
    if match:
        status, value_str, unit = match.groups()
        value = float(value_str.replace(' ', ''))
        return status, value, unit
    return None, None, None

ohaus_response_pattern = re.compile(r"([+-]?\s*[\d\.]+)\s+([a-zA-Z]+)")
def parse_ohaus_response(response_str):
    match = ohaus_response_pattern.search(response_str)
    if match:
        val_str, unit = match.groups()
        value = float(val_str.replace(' ', ''))
        return "S", value, unit
    return None, None, None

# --- Serial Worker Thread ---
class SerialWorker(QThread):
    data_received = pyqtSignal(float, str)
    
    def __init__(self, port, baudrate, brand):
        super().__init__()
        self.port = port
        self.baudrate = baudrate
        self.brand = brand
        self.running = True
        self.ser = None
        
        try:
            self.ser = serial.Serial(
                port=self.port, baudrate=self.baudrate, timeout=0.1,
                bytesize=serial.EIGHTBITS, parity=serial.PARITY_NONE, stopbits=serial.STOPBITS_ONE,
                xonxoff=False, rtscts=False, dsrdtr=False
            )
        except Exception as e:
            self.error_msg = str(e)
            self.running = False

    def run(self):
        while self.running and self.ser and self.ser.is_open:
            try:
                if self.brand == "Mettler Toledo":
                    if self.ser.in_waiting == 0:
                        self.ser.write(b"SI\r\n")
                elif self.brand == "Ohaus Adventurer":
                    if self.ser.in_waiting == 0:
                        self.ser.write(b"IP\r\n")
                        
                response = self.ser.readline()
                if response:
                    decoded = response.decode('ascii', errors='ignore').strip()
                    if not decoded: continue
                        
                    if self.brand == "Mettler Toledo":
                        status, value, unit = parse_mt_sics_response(decoded)
                    elif self.brand == "Ohaus Adventurer":
                        status, value, unit = parse_ohaus_response(decoded)
                    else:
                        status, value, unit = parse_bonvoisin_response(decoded)
                        
                    if value is not None:
                        self.data_received.emit(value, unit)
                        
            except Exception:
                time.sleep(0.1)

    def stop(self):
        self.running = False
        if self.ser and self.ser.is_open:
            try:
                self.ser.close()
            except: pass

    def send_tare(self):
        if self.ser and self.ser.is_open:
            try:
                self.ser.write(b"T\r\n")
            except: pass



class CollapsibleBox(QWidget):
    def __init__(self, title, parent=None):
        super().__init__(parent)
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 0, 0, 0)
        self.layout.setSpacing(0)
        
        self.btn_toggle = QPushButton(f"▼ {title}")
        self.btn_toggle.setStyleSheet("text-align: left; font-weight: bold; padding: 5px; background-color: transparent; border: none;")
        self.btn_toggle.setCheckable(True)
        self.btn_toggle.setChecked(False)  # False = expanded
        self.btn_toggle.toggled.connect(self.on_toggle)
        self.layout.addWidget(self.btn_toggle)
        
        self.content_widget = QWidget()
        self.content_layout = QVBoxLayout(self.content_widget)
        self.content_layout.setContentsMargins(5, 5, 5, 5)
        self.layout.addWidget(self.content_widget)
        
        self.title_text = title
        
    def on_toggle(self, checked):
        self.btn_toggle.setText(f"▶ {self.title_text}" if checked else f"▼ {self.title_text}")
        self.content_widget.setVisible(not checked)

    def set_collapsed(self, collapsed=True):
        self.btn_toggle.setChecked(collapsed)

class ReorderPanelsWidget(QWidget):
    def __init__(self, app_ref, parent=None):
        super().__init__(parent)
        self.app_ref = app_ref
        
        
        
        layout = QVBoxLayout(self)
        
        lbl = QLabel("Drag and drop to reorder panels. The app will restart when applied.")
        layout.addWidget(lbl)
        
        lists_layout = QHBoxLayout()
        
        left_grp = QGroupBox("Left Panel")
        left_lay = QVBoxLayout(left_grp)
        self.list_left = QListWidget()
        self.list_left.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        left_order = self.app_ref.config.get("left_panel_order", [
            "Recording Settings", "Derivative Engine", "Secondary Smoothing", 
            "Pump Calibration", "Axis Limits", "Data Analysis"
        ])
        self.list_left.addItems(left_order)
        left_lay.addWidget(self.list_left)
        lists_layout.addWidget(left_grp)
        
        right_grp = QGroupBox("Right Panel")
        right_lay = QVBoxLayout(right_grp)
        self.list_right = QListWidget()
        self.list_right.setDragDropMode(QAbstractItemView.DragDropMode.InternalMove)
        right_order = self.app_ref.config.get("right_panel_order", [
            "Tab Management", "Connection", "Experiment Notes"
        ])
        self.list_right.addItems(right_order)
        right_lay.addWidget(self.list_right)
        lists_layout.addWidget(right_grp)
        
        layout.addLayout(lists_layout)
        
        btn_layout = QHBoxLayout()
        btn_apply = QPushButton("Apply and Restart")
        btn_apply.clicked.connect(self.apply)
        
        
        btn_layout.addStretch()
        btn_layout.addWidget(btn_apply)
        layout.addLayout(btn_layout)
        
    def apply(self):
        new_left = [self.list_left.item(i).text() for i in range(self.list_left.count())]
        new_right = [self.list_right.item(i).text() for i in range(self.list_right.count())]
        self.app_ref.config["left_panel_order"] = new_left
        self.app_ref.config["right_panel_order"] = new_right
        self.app_ref.save_config()
        pass
        
        # Restart the app
        import sys
        from PyQt6.QtCore import QProcess
        from PyQt6.QtWidgets import QApplication
        
        QProcess.startDetached(sys.executable, sys.argv)
        QApplication.instance().quit()
class GlobalSavingSettingsWidget(QWidget):
    def __init__(self, parent=None, config=None):
        super().__init__(parent)
        
        
        
        self.config = config or {}
        
        main_layout = QVBoxLayout(self)
        
        self.tabs = QTabWidget()
        main_layout.addWidget(self.tabs)
        
        # --- TAB 1: General Settings ---
        tab_general = QWidget()
        l_general = QVBoxLayout(tab_general)
        
        # Save Folder
        gb_folder = QGroupBox("Global Default Save Folder")
        l_folder = QHBoxLayout(gb_folder)
        self.ent_folder = QLineEdit(self.config.get("global_default_save_folder", "Data"))
        btn_browse = QPushButton("Browse")
        btn_browse.clicked.connect(self.browse_folder)
        l_folder.addWidget(self.ent_folder)
        l_folder.addWidget(btn_browse)
        l_general.addWidget(gb_folder)
        
        # Template
        gb_temp = QGroupBox("Global Filename Template")
        l_temp = QVBoxLayout(gb_temp)
        
        form_layout = QFormLayout()
        self.ent_template = QLineEdit(self.config.get("global_filename_template", "<Tab Name>_<Date>_<Time>"))
        form_layout.addRow("Filename Format:", self.ent_template)
        l_temp.addLayout(form_layout)
        
        l_temp.addWidget(QLabel("Click a token below to insert it at the cursor position:"))
        
        self.grid_buttons = QGridLayout()
        l_temp.addLayout(self.grid_buttons)
        l_general.addWidget(gb_temp)
        l_general.addStretch()
        
        self.tabs.addTab(tab_general, "General & Template")
        
        # --- TAB 2: Metadata Tokens ---
        tab_tokens = QWidget()
        l_tokens_wrapper = QVBoxLayout(tab_tokens)
        
        scroll_tokens = QScrollArea()
        scroll_tokens.setWidgetResizable(True)
        scroll_content = QWidget()
        
        self.layout_tokens = QVBoxLayout(scroll_content)
        self.token_rows = []
        self.token_button_refs = []
        
        btn_add_token = QPushButton("+ Add Custom Token")
        btn_add_token.clicked.connect(lambda: self.add_token_row("", ""))
        self.layout_tokens.addWidget(btn_add_token)
        
        global_tokens = self.config.get("global_tokens", [
            {"name": "Column Name", "value": ""},
            {"name": "Material", "value": ""},
            {"name": "Flow rate", "value": ""},
            {"name": "Bed length", "value": ""},
            {"name": "Run #", "value": ""}
        ])
        for t in global_tokens:
            self.add_token_row(t.get("name", ""), t.get("value", ""))
            
        scroll_tokens.setWidget(scroll_content)
        l_tokens_wrapper.addWidget(scroll_tokens)
        
        self.tabs.addTab(tab_tokens, "Metadata Tokens")
        self.refresh_token_buttons()
        
        # Buttons
        btn_layout = QHBoxLayout()
        btn_ok = QPushButton("Save Globally")
        
        btn_layout.addWidget(btn_ok)
        main_layout.addLayout(btn_layout)
        
        btn_ok.clicked.connect(self.apply_settings)
        
        
    def apply_settings(self):
        self.config["global_default_save_folder"] = self.ent_folder.text()
        self.config["global_filename_template"] = self.ent_template.text()
        self.config["global_tokens"] = self.get_tokens()
        if hasattr(self.parent(), 'app_ref'):
            self.parent().app_ref.save_config()
            for tab in self.parent().app_ref.tab_objects:
                if hasattr(tab, 'refresh_tokens'):
                    tab.refresh_tokens()

    def browse_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Save Folder", self.ent_folder.text())
        if folder: self.ent_folder.setText(folder)
        
    def add_token_row(self, name, value):
        row_widget = QWidget()
        row_layout = QHBoxLayout(row_widget)
        row_layout.setContentsMargins(0, 0, 0, 0)
        
        ent_name = QLineEdit(name)
        ent_name.setPlaceholderText("Token Name")
        ent_name.textChanged.connect(self.refresh_token_buttons)
        ent_value = QLineEdit(value)
        ent_value.setPlaceholderText("Value")
        
        btn_remove = QPushButton("X")
        btn_remove.setFixedWidth(30)
        btn_remove.setStyleSheet("background-color: #e74c3c; color: white;")
        
        row_layout.addWidget(ent_name)
        row_layout.addWidget(ent_value)
        row_layout.addWidget(btn_remove)
        
        idx = self.layout_tokens.count() - 1
        self.layout_tokens.insertWidget(idx, row_widget)
        
        row_data = (row_widget, ent_name, ent_value)
        self.token_rows.append(row_data)
        
        btn_remove.clicked.connect(lambda: self.remove_token_row(row_data))
        self.refresh_token_buttons()
        
    def remove_token_row(self, row_data):
        row_widget, ent_name, ent_value = row_data
        if row_data in self.token_rows:
            self.token_rows.remove(row_data)
        self.layout_tokens.removeWidget(row_widget)
        row_widget.hide()
        row_widget.setParent(None)
        row_widget.deleteLater()
        self.refresh_token_buttons()
        
    def refresh_token_buttons(self):
        while self.grid_buttons.count():
            item = self.grid_buttons.takeAt(0)
            if item.widget():
                item.widget().hide()
                item.widget().setParent(None)
                item.widget().deleteLater()
        
        self.token_button_refs = []
            
        tokens = ["Tab Name", "Date", "Time"]
        for _, ent_name, _ in self.token_rows:
            n = ent_name.text().strip()
            if n and n not in tokens: tokens.append(n)
            
        row = 0; col = 0
        for tok in tokens:
            btn = QPushButton(f"<{tok}>")
            btn.clicked.connect(lambda checked, t=tok: self.ent_template.insert(f"<{t}>"))
            self.grid_buttons.addWidget(btn, row, col)
            self.token_button_refs.append(btn)
            col += 1
            if col > 3:
                col = 0
                row += 1
                
    def get_tokens(self):
        return [{"name": n.text(), "value": v.text()} for _, n, v in self.token_rows if n.text().strip()]


class SettingsTab(QWidget):
    def __init__(self, app_ref, parent=None):
        super().__init__(parent)
        self.app_ref = app_ref
        self.tab_name = "Settings"
        
        main_layout = QVBoxLayout(self)
        
        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll_content = QWidget()
        scroll_layout = QVBoxLayout(scroll_content)
        
        # 1. Application Settings
        gb_app = QGroupBox("Application Settings")
        l_app = QVBoxLayout(gb_app)
        
        self.chk_btn_text = QCheckBox("Show Button Text")
        self.chk_btn_text.setChecked(self.app_ref.config.get("show_button_text", False))
        self.chk_btn_text.clicked.connect(self.toggle_button_text)
        l_app.addWidget(self.chk_btn_text)
        
        self.chk_dark_mode = QCheckBox("Dark Mode")
        self.chk_dark_mode.setChecked(self.app_ref.config.get("dark_mode", True))
        self.chk_dark_mode.clicked.connect(self.toggle_theme)
        l_app.addWidget(self.chk_dark_mode)
        
        btn_restart = QPushButton("Restart Application")
        btn_restart.clicked.connect(self.app_ref.restart_app)
        l_app.addWidget(btn_restart)
        scroll_layout.addWidget(gb_app)
        
        # 2. Panel Management
        gb_panels = QGroupBox("Panel Management")
        l_panels = QVBoxLayout(gb_panels)
        self.panel_widget = ReorderPanelsWidget(self.app_ref)
        l_panels.addWidget(self.panel_widget)
        scroll_layout.addWidget(gb_panels)
        
        # 3. Global Settings & Metadata
        gb_global = QGroupBox("Global Settings & Metadata")
        l_global = QVBoxLayout(gb_global)
        self.global_widget = GlobalSavingSettingsWidget(self, self.app_ref.config)
        l_global.addWidget(self.global_widget)
        scroll_layout.addWidget(gb_global)
        
        # 4. Help
        gb_help = QGroupBox("Help")
        l_help = QVBoxLayout(gb_help)
        btn_help = QPushButton("About Flow Rate Engine")
        btn_help.clicked.connect(self.app_ref.show_help_dialog)
        l_help.addWidget(btn_help)
        scroll_layout.addWidget(gb_help)
        
        scroll_layout.addStretch()
        
        scroll.setWidget(scroll_content)
        main_layout.addWidget(scroll)
        
    def toggle_button_text(self, checked):
        self.app_ref.config["show_button_text"] = checked
        self.app_ref.save_config()
        for tab in self.app_ref.tab_objects:
            if hasattr(tab, 'update_gui_components'):
                tab.update_gui_components()
                
    def toggle_theme(self, checked):
        self.app_ref.config["dark_mode"] = checked
        self.app_ref.save_config()
        self.app_ref.apply_theme()

class BalanceTab(QWidget):
    def __init__(self, parent=None, app=None, tab_name="Balance"):
        super().__init__(parent)
        self.app = app
        self.tab_name = tab_name
        self.serial_thread = None
        
        # Data tracking
        self.times_sec = []
        self.times_min = []
        self.weights = []
        self.flow_rates = []
        self.timestamps = []
        self.start_time = None
        self.last_recorded_time = 0
        self.unsaved_changes = False
        
        self.recording = False
        self.current_unit = "g"
        self.unit_set = False
        
        # Backup path
        self.setup_backup_path()
        
        self.build_ui()
        self.load_tab_settings()
        
        # Update timer for GUI plotting
        self.update_timer = QTimer()
        self.update_timer.timeout.connect(self.update_gui_components)
        self.update_timer.start(500) # 2 Hz update

    def setup_backup_path(self):
        backup_dir = os.path.join(os.getcwd(), "Data Backups")
        os.makedirs(backup_dir, exist_ok=True)
        self.backup_path = os.path.join(backup_dir, f"{self.tab_name.replace(' ', '_')}_Backup.csv")
        self.prune_backup_file()

    def update_button_text(self):
        show_text = self.app.config.get("show_button_text", False) if self.app else False
        
        if self.recording:
            self.btn_record.setText("⏹ Stop Recording" if show_text else "⏹")
        else:
            self.btn_record.setText("▶ Start Recording" if show_text else "▶")
            
        self.btn_clear.setText("🗑 Clear Data" if show_text else "🗑")
        self.btn_save_excel.setText("📊 Save to Excel" if show_text else "📊")
        self.btn_tare.setText("⚖ Tare Balance" if show_text else "⚖")
        self.btn_save_graph.setText("🖼 Save Graph PNG" if show_text else "🖼")

    def build_ui(self):
        main_layout = QVBoxLayout(self)
        
        self.top_action_layout = QHBoxLayout()
        
        self.btn_record = QPushButton()
        self.btn_record.setStyleSheet("background-color: #2ecc71; color: white; font-weight: bold;")
        self.btn_record.clicked.connect(self.toggle_record)
        self.btn_record.setToolTip("Start Recording")
        
        self.btn_clear = QPushButton()
        self.btn_clear.setStyleSheet("background-color: #e74c3c; color: white;")
        self.btn_clear.clicked.connect(self.clear_data)
        self.btn_clear.setToolTip("Clear Data")
        
        self.btn_save_excel = QPushButton()
        self.btn_save_excel.setStyleSheet("background-color: #27ae60; color: white; font-weight: bold;")
        self.btn_save_excel.clicked.connect(self.save_excel)
        self.btn_save_excel.setToolTip("Save to Excel")
        
        self.btn_tare = QPushButton()
        self.btn_tare.clicked.connect(self.tare_balance)
        self.btn_tare.setToolTip("Tare Balance")
        
        self.btn_save_graph = QPushButton()
        self.btn_save_graph.setStyleSheet("background-color: #2980b9; color: white;")
        self.btn_save_graph.clicked.connect(self.save_graph)
        self.btn_save_graph.setToolTip("Save Graph PNG")
        
        self.top_action_layout.addWidget(self.btn_record)
        self.top_action_layout.addWidget(self.btn_clear)
        self.top_action_layout.addWidget(self.btn_save_excel)
        self.top_action_layout.addWidget(self.btn_tare)
        self.top_action_layout.addWidget(self.btn_save_graph)
        self.top_action_layout.addStretch()
        
        main_layout.addLayout(self.top_action_layout)
        
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        main_layout.addWidget(self.splitter)
        
        self.update_button_text()
        
        # --- LEFT PANEL ---
        left_widget = QWidget()
        left_layout = QVBoxLayout(left_widget)
        left_scroll = QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setWidget(left_widget)
        
        def create_form_row(layout, label, default=""):
            row = QHBoxLayout()
            lbl = QLabel(label)
            ent = QLineEdit(default)
            row.addWidget(lbl)
            row.addWidget(ent)
            layout.addLayout(row)
            return ent

        self.left_panels = {}

        # Recording Settings
        gb_rec = CollapsibleBox("Recording Settings")
        l_rec = gb_rec.content_layout
        self.ent_interval = create_form_row(l_rec, "Interval (s):", "1.0")
        
        self.chk_auto_stop = QCheckBox("Enable Auto-Stop")
        l_rec.addWidget(self.chk_auto_stop)
        self.ent_auto_stop_min = create_form_row(l_rec, "Auto-Stop After (min):", "5.0")
        self.ent_auto_stop_thresh = create_form_row(l_rec, "Flow Threshold:", "0.1")
        self.left_panels["Recording Settings"] = gb_rec
        
        # Derivative Engine
        gb_der = CollapsibleBox("Derivative Engine")
        l_der = gb_der.content_layout
        self.ent_savgol_win = create_form_row(l_der, "SavGol Window:", "5")
        self.ent_savgol_poly = create_form_row(l_der, "SavGol Poly:", "3")
        self.left_panels["Derivative Engine"] = gb_der
        
        # Filtering
        gb_filt = CollapsibleBox("Secondary Smoothing")
        l_filt = gb_filt.content_layout
        self.combo_filter = QComboBox()
        self.combo_filter.addItems(["Mean", "Median", "EMA", "Butterworth", "Adaptive"])
        self.combo_filter.currentTextChanged.connect(self.on_filter_change)
        l_filt.addWidget(self.combo_filter)
        
        self.lbl_filter_param = QLabel("Window Size (samples):")
        l_filt.addWidget(self.lbl_filter_param)
        self.ent_filter_param = QLineEdit("20")
        l_filt.addWidget(self.ent_filter_param)
        
        self.btn_apply_filter = QPushButton("Apply Smoothing")
        self.btn_apply_filter.setStyleSheet("background-color: #27ae60; color: white;")
        self.btn_apply_filter.clicked.connect(self.manual_apply_filters)
        l_filt.addWidget(self.btn_apply_filter)
        
        self.left_panels["Secondary Smoothing"] = gb_filt
        
        # Calibration
        gb_cal = CollapsibleBox("Pump Calibration")
        gb_cal.set_collapsed(True)
        l_cal = gb_cal.content_layout
        self.ent_rpm = create_form_row(l_cal, "Pump RPM:", "30")
        self.ent_rollers = create_form_row(l_cal, "Rollers:", "3")
        self.btn_log_cal = QPushButton("Log Calibration Point")
        self.btn_log_cal.setStyleSheet("background-color: #8e44ad; color: white; font-weight: bold;")
        self.btn_log_cal.clicked.connect(self.log_calibration)
        l_cal.addWidget(self.btn_log_cal)
        self.left_panels["Pump Calibration"] = gb_cal
        
        # Axis Limits
        gb_axis = CollapsibleBox("Axis Limits")
        gb_axis.set_collapsed(True)
        l_axis = gb_axis.content_layout
        
        self.chk_auto_x = QCheckBox("Auto Time")
        self.chk_auto_x.setChecked(True)
        self.chk_auto_y_mass = QCheckBox("Auto Mass")
        self.chk_auto_y_mass.setChecked(True)
        self.chk_auto_y_flow = QCheckBox("Auto Flow")
        self.chk_auto_y_flow.setChecked(True)
        self.chk_show_flow = QCheckBox("Show Flow Rate")
        self.chk_show_flow.setChecked(True)
        self.chk_show_flow.stateChanged.connect(self.toggle_flow_axis)
        
        l_axis.addWidget(self.chk_auto_x)
        l_axis.addWidget(self.chk_auto_y_mass)
        l_axis.addWidget(self.chk_auto_y_flow)
        l_axis.addWidget(self.chk_show_flow)
        
        self.combo_xaxis = QComboBox()
        self.combo_xaxis.addItems(["Duration", "Timestamp"])
        l_axis.addWidget(QLabel("Plot X-Axis By:"))
        l_axis.addWidget(self.combo_xaxis)
        
        self.ent_xmin = create_form_row(l_axis, "X Min (min):", "0")
        self.ent_xmax = create_form_row(l_axis, "X Max (min):", "10")
        self.ent_ymin = create_form_row(l_axis, "Mass Min (g):", "0")
        self.ent_ymax = create_form_row(l_axis, "Mass Max (g):", "100")
        self.ent_flowmin = create_form_row(l_axis, "Flow Min:", "0")
        self.ent_flowmax = create_form_row(l_axis, "Flow Max:", "20")
        
        self.btn_apply_lims = QPushButton("Apply Manual Limits")
        self.btn_apply_lims.clicked.connect(self.apply_axis_limits)
        l_axis.addWidget(self.btn_apply_lims)
        self.left_panels["Axis Limits"] = gb_axis
        
        # Data Analysis
        gb_ana = CollapsibleBox("Data Analysis")
        l_ana = gb_ana.content_layout
        self.ent_fit_start = create_form_row(l_ana, "Fit Start (min):", "0")
        self.ent_fit_end = create_form_row(l_ana, "Fit End (min):", "10")
        self.btn_apply_fit = QPushButton("Apply Linear Fit")
        self.btn_apply_fit.setStyleSheet("background-color: #e67e22; color: white;")
        self.btn_apply_fit.clicked.connect(self.apply_linear_fit)
        l_ana.addWidget(self.btn_apply_fit)
        self.left_panels["Data Analysis"] = gb_ana
        
        left_order = self.app.config.get("left_panel_order", [
            "Recording Settings", "Derivative Engine", "Secondary Smoothing", 
            "Pump Calibration", "Axis Limits", "Data Analysis"
        ]) if self.app else ["Recording Settings", "Derivative Engine", "Secondary Smoothing", "Pump Calibration", "Axis Limits", "Data Analysis"]

        for name in left_order:
            if name in self.left_panels:
                left_layout.addWidget(self.left_panels[name])
                
        left_layout.addStretch()
        
        self.splitter.addWidget(left_scroll)
        
        # --- MIDDLE PANEL ---
        mid_widget = QWidget()
        mid_layout = QVBoxLayout(mid_widget)
        self.mid_splitter = QSplitter(Qt.Orientation.Vertical)
        mid_layout.addWidget(self.mid_splitter)
        
        # Plot
        self.fig, self.ax_mass = plt.subplots(figsize=(8, 6))
        self.fig.patch.set_facecolor('white')
        self.ax_mass.set_facecolor('white')
        self.ax_flow = self.ax_mass.twinx()
        
        self.ax_mass.set_title(f"{self.tab_name} - Live Weight Data", color='black')
        self.ax_mass.set_xlabel("Time (minutes)", color='black')
        self.ax_mass.set_ylabel("Weight (g)", color='#2980b9')
        self.ax_mass.tick_params(axis='x', colors='black')
        self.ax_mass.tick_params(axis='y', labelcolor='#2980b9')
        self.ax_mass.grid(True, alpha=0.3, color='gray')
        
        self.ax_flow.set_ylabel("Flow Rate (g/min)", color='#c0392b')
        self.ax_flow.tick_params(axis='y', labelcolor='#c0392b')
        
        self.line_mass, = self.ax_mass.plot([], [], marker='o', linestyle='-', color='#3498db', label='Mass')
        self.line_flow, = self.ax_flow.plot([], [], marker='', linestyle='-', color='#e74c3c', label='Flow')
        self.line_fit, = self.ax_mass.plot([], [], marker='', linestyle='--', color='#f1c40f', label='Linear Fit', linewidth=2.5)
        self.fig.tight_layout()
        
        self.canvas = FigureCanvasQTAgg(self.fig)
        self.mid_splitter.addWidget(self.canvas)
        
        # Table
        self.table = QTableWidget()
        self.table.setColumnCount(4)
        self.table.setHorizontalHeaderLabels(["Timestamp", "Duration (min)", "Mass (g)", "Flow (g/min)"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.table.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self.mid_splitter.addWidget(self.table)
        
        self.splitter.addWidget(mid_widget)
        
        # --- RIGHT PANEL ---
        right_widget = QWidget()
        right_layout = QVBoxLayout(right_widget)
        right_scroll = QScrollArea()
        right_scroll.setWidgetResizable(True)
        right_scroll.setWidget(right_widget)
        
        # Tab Management
        gb_tab = QGroupBox("Tab Management")
        l_tab = QVBoxLayout(gb_tab)
        
        l_tab.addWidget(QLabel("New Tab Name:"))
        self.ent_new_tab = QLineEdit()
        l_tab.addWidget(self.ent_new_tab)
        
        btn_add_tab = QPushButton("Add Tab")
        btn_add_tab.clicked.connect(self.request_add_tab)
        l_tab.addWidget(btn_add_tab)
        
        btn_rename_tab = QPushButton("Rename Current Tab")
        btn_rename_tab.clicked.connect(self.request_rename_tab)
        l_tab.addWidget(btn_rename_tab)
        
        self.btn_save_tab = QPushButton("Save Tab Configuration")
        self.btn_save_tab.clicked.connect(self.save_tab_config)
        l_tab.addWidget(self.btn_save_tab)
        
        self.btn_load_tab = QPushButton("Load Tab Configuration")
        self.btn_load_tab.clicked.connect(self.load_tab_config_dialog)
        l_tab.addWidget(self.btn_load_tab)
        
        self.btn_close_tab = QPushButton("✖ Close Tab")
        self.btn_close_tab.setStyleSheet("background-color: #e74c3c; color: white;")
        self.btn_close_tab.clicked.connect(self.close_tab)
        l_tab.addWidget(self.btn_close_tab)
        
        right_layout.addWidget(gb_tab)
        
        # Connection
        gb_conn = QGroupBox("Connection")
        l_conn = QVBoxLayout(gb_conn)
        l_conn.addWidget(QLabel("Balance Brand:"))
        self.combo_brand = QComboBox()
        self.combo_brand.addItems(["Bonvoisin", "Mettler Toledo", "Ohaus Adventurer", "Lachoi"])
        l_conn.addWidget(self.combo_brand)
        
        l_conn.addWidget(QLabel("COM Port:"))
        self.combo_com = QComboBox()
        self.refresh_com_ports()
        l_conn.addWidget(self.combo_com)
        
        self.btn_refresh = QPushButton("Refresh Ports")
        self.btn_refresh.clicked.connect(self.refresh_com_ports)
        l_conn.addWidget(self.btn_refresh)
        
        self.btn_connect = QPushButton("Connect")
        self.btn_connect.setStyleSheet("background-color: #3498db; color: white; font-weight: bold;")
        self.btn_connect.clicked.connect(self.connect_serial)
        l_conn.addWidget(self.btn_connect)
        
        self.btn_disconnect = QPushButton("Disconnect")
        self.btn_disconnect.setStyleSheet("background-color: #e74c3c; color: white;")
        self.btn_disconnect.clicked.connect(self.disconnect_serial)
        l_conn.addWidget(self.btn_disconnect)
        
        self.lbl_status = QLabel("Status: Disconnected")
        self.lbl_status.setStyleSheet("color: gray; font-weight: bold;")
        l_conn.addWidget(self.lbl_status)
        right_layout.addWidget(gb_conn)
        
        # Save Settings
        gb_save = QGroupBox("Auto-Save Settings")
        l_save = QVBoxLayout(gb_save)
        
        l_save.addWidget(QLabel("Save Folder Path:"))
        folder_layout = QHBoxLayout()
        default_folder = self.app.config.get("global_default_save_folder", "Data") if self.app else "Data"
        self.ent_save_folder = QLineEdit(default_folder)
        self.ent_save_folder.setMinimumHeight(35)
        btn_browse = QPushButton("Browse")
        btn_browse.setMinimumHeight(35)
        btn_browse.clicked.connect(self.browse_save_folder)
        folder_layout.addSpacing(20)
        folder_layout.addWidget(self.ent_save_folder)
        folder_layout.addWidget(btn_browse)
        l_save.addLayout(folder_layout)
        
        l_save.addSpacing(10)
        
        # Template
        self.chk_use_global_template = QCheckBox("Use Global Filename Template")
        self.chk_use_global_template.setChecked(True)
        self.chk_use_global_template.stateChanged.connect(self.toggle_template_source)
        l_save.addWidget(self.chk_use_global_template)
        
        temp_layout = QHBoxLayout()
        temp_layout.addWidget(QLabel("Tab Template:"))
        self.ent_save_template = QLineEdit()
        if self.app:
            self.ent_save_template.setText(self.app.config.get("global_filename_template", "<Tab Name>_<Date>_<Time>"))
        self.ent_save_template.setEnabled(False)
        temp_layout.addWidget(self.ent_save_template)
        l_save.addLayout(temp_layout)
        
        l_save.addSpacing(10)
        
        mode_layout = QFormLayout()
        self.combo_autosave = QComboBox()
        self.combo_autosave.addItems(["Off", "On Auto-Stop", "Every X Minutes"])
        mode_layout.addRow("Auto-Save Mode:", self.combo_autosave)
        
        self.ent_autosave_min = QLineEdit("5.0")
        mode_layout.addRow("Auto-Save Interval (min):", self.ent_autosave_min)
        l_save.addLayout(mode_layout)
        
        self.combo_autosave.currentIndexChanged.connect(self.toggle_autosave_interval)
        self.toggle_autosave_interval()
        
        right_layout.addWidget(gb_save)
        
        # Metadata Tokens
        self.gb_tokens = QGroupBox("Metadata Tokens")
        self.tokens_outer_layout = QVBoxLayout(self.gb_tokens)
        self.l_tokens = QFormLayout()
        self.tokens_outer_layout.addLayout(self.l_tokens)
        
        self.btn_add_tab_token = QPushButton("+ Add Custom Token")
        self.btn_add_tab_token.clicked.connect(self.request_add_custom_token)
        self.tokens_outer_layout.addWidget(self.btn_add_tab_token)
        
        self.token_inputs = {}
        right_layout.addWidget(self.gb_tokens)
        self.refresh_tokens()
        
        # Notes
        gb_notes = QGroupBox("Experiment Notes")
        l_notes = QVBoxLayout(gb_notes)
        self.txt_notes = QTextEdit()
        self.txt_notes.setMinimumHeight(150)
        l_notes.addWidget(self.txt_notes)
        right_layout.addWidget(gb_notes)
        
        self.btn_recover = QPushButton("Recover Session")
        self.btn_recover.setStyleSheet("background-color: #8e44ad; color: white;")
        self.btn_recover.clicked.connect(self.prompt_session_recovery)
        right_layout.addWidget(self.btn_recover)
        
        # btn_save_tab moved to Tab Management
        
        right_layout.addStretch()
        self.splitter.addWidget(right_scroll)
        
        self.splitter.setSizes([300, 600, 300])

    def refresh_tokens(self):
        old_values = {k: v.text() for k, v in getattr(self, 'token_inputs', {}).items()}
        
        while self.l_tokens.rowCount() > 0:
            self.l_tokens.removeRow(0)
            
        self.token_inputs = {}
        
        global_tokens = self.app.config.get("global_tokens", []) if self.app else []
        global_names = []
        for t in global_tokens:
            name = t.get("name", "").strip()
            if name:
                global_names.append(name)
                ent = QLineEdit()
                if name in old_values:
                    ent.setText(old_values[name])
                else:
                    ent.setText(t.get("value", ""))
                self.l_tokens.addRow(name + ":", ent)
                self.token_inputs[name] = ent
                
        for name, val in old_values.items():
            if name not in global_names:
                self.add_custom_token_ui(name, val)

    def request_add_custom_token(self):
        from PyQt6.QtWidgets import QInputDialog
        name, ok = QInputDialog.getText(self, "Add Custom Token", "Token Name:")
        if ok and name.strip():
            name = name.strip()
            if name in self.token_inputs:
                QMessageBox.warning(self, "Warning", "Token already exists!")
                return
            self.add_custom_token_ui(name, "")

    def add_custom_token_ui(self, name, value):
        ent = QLineEdit(value)
        row_layout = QHBoxLayout()
        row_layout.addWidget(ent)
        btn_del = QPushButton("-")
        btn_del.setFixedWidth(20)
        btn_del.clicked.connect(lambda _, n=name: self.remove_custom_token(n))
        row_layout.addWidget(btn_del)
        self.l_tokens.addRow(name + ":", row_layout)
        self.token_inputs[name] = ent

    def remove_custom_token(self, name):
        if name in self.token_inputs:
            self.token_inputs.pop(name)
            self.refresh_tokens()

    def refresh_com_ports(self):
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self.combo_com.clear()
        if ports:
            self.combo_com.addItems(ports)
        else:
            self.combo_com.addItem("No Ports Found")

    def toggle_flow_axis(self):
        self.ax_flow.set_visible(self.chk_show_flow.isChecked())
        self.canvas.draw()

    def on_filter_change(self, choice):
        if choice in ["Mean", "Median"]:
            self.lbl_filter_param.setText("Window Size (samples):")
            if choice == "Mean": self.ent_filter_param.setText("20")
        elif choice == "EMA":
            self.lbl_filter_param.setText("Alpha (0.01 - 1.0):")
            self.ent_filter_param.setText("0.1")
        elif choice == "Butterworth":
            self.lbl_filter_param.setText("Cutoff Freq (Hz):")
            self.ent_filter_param.setText("0.01")
        elif choice == "Adaptive":
            self.lbl_filter_param.setText("Window (Auto-calculated)")

    def connect_serial(self, auto=False):
        port = self.combo_com.currentText()
        if not port or port == "No Ports Found":
            if not auto: QMessageBox.critical(self, "Error", "Invalid COM Port.")
            return
            
        if self.serial_thread and self.serial_thread.isRunning():
            self.serial_thread.stop()
            self.serial_thread.wait()
            
        baud = 9600
        brand = self.combo_brand.currentText()
        
        self.serial_thread = SerialWorker(port, baud, brand)
        if hasattr(self.serial_thread, 'error_msg'):
            self.lbl_status.setText("Status: Connection Failed")
            self.lbl_status.setStyleSheet("color: #e74c3c;")
            if not auto: QMessageBox.critical(self, "Error", f"Could not connect:\n{self.serial_thread.error_msg}")
            return
            
        self.serial_thread.data_received.connect(self.on_data_received)
        self.serial_thread.start()
        
        self.lbl_status.setText(f"Status: Connected ({port})")
        self.lbl_status.setStyleSheet("color: #2ecc71; font-weight: bold;")
        self.btn_connect.setText("Connected")
        self.btn_connect.setStyleSheet("background-color: #2ecc71; color: white;")
        
        if self.app:
            self.app.save_connection(self.tab_name, brand, port)

    def disconnect_serial(self):
        if self.serial_thread:
            self.serial_thread.stop()
            self.serial_thread.wait()
            self.serial_thread = None
        self.lbl_status.setText("Status: Disconnected")
        self.lbl_status.setStyleSheet("color: gray;")
        self.btn_connect.setText("Connect")
        self.btn_connect.setStyleSheet("background-color: #3498db; color: white; font-weight: bold;")

    def on_data_received(self, value, unit):
        if not self.recording:
            return
            
        try:
            intrvl = float(self.ent_interval.text())
        except: intrvl = 1.0
        
        curr = time.time() - self.start_time
        if (curr - self.last_recorded_time >= intrvl) or len(self.times_sec) == 0:
            self.last_recorded_time = curr
            
            self.times_sec.append(curr)
            self.times_min.append(curr / 60.0)
            self.weights.append(value)
            
            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S.%f")[:-3]
            self.timestamps.append(ts)
            
            try:
                is_new = not os.path.exists(self.backup_path)
                with open(self.backup_path, 'a', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    if is_new:
                        writer.writerow(["Timestamp", "Duration (min)", "Mass (g)", "Flow Rate (g/min)"])
                    writer.writerow([ts, round(curr / 60.0, 5), value, ""])
            except Exception as e:
                pass
                
            self.unsaved_changes = True
            if self.app: self.app.set_unsaved_state(self.tab_name, True)
            
            if not self.unit_set:
                self.current_unit = unit
                self.unit_set = True

    def calculate_flows(self):
        n = len(self.times_sec)
        while len(self.flow_rates) < n:
            self.flow_rates.append(None)
            
        valid_indices = []
        raw_flows = [None] * n
        
        if n >= 5:
            try:
                sg_win = max(5, int(self.ent_savgol_win.text()))
                if sg_win % 2 == 0: sg_win += 1
                sg_poly = min(sg_win - 1, int(self.ent_savgol_poly.text()))
                
                import numpy as np
                import scipy.signal
                
                if n >= sg_win:
                    w_sm = scipy.signal.savgol_filter(self.weights, window_length=sg_win, polyorder=sg_poly)
                    deriv_raw = np.gradient(w_sm, self.times_sec)
                else:
                    deriv_raw = np.gradient(self.weights, self.times_sec)
                    
                for i in range(n):
                    raw_flows[i] = deriv_raw[i] * 60.0
                    valid_indices.append(i)
            except:
                import numpy as np
                deriv_raw = np.gradient(self.weights, self.times_sec)
                for i in range(n):
                    raw_flows[i] = deriv_raw[i] * 60.0
                    valid_indices.append(i)
        elif n > 1:
            import numpy as np
            deriv_raw = np.gradient(self.weights, self.times_sec)
            for i in range(n):
                raw_flows[i] = deriv_raw[i] * 60.0
                valid_indices.append(i)
                
        if not valid_indices: return
        
        filter_type = self.combo_filter.currentText()
        
        if filter_type in ['Mean', 'Median']:
            try: win = max(1, int(self.ent_filter_param.text()))
            except: win = 20
            
            for v_idx in range(len(valid_indices)):
                if v_idx >= win - 1:
                    i = valid_indices[v_idx]
                    window_indices = valid_indices[v_idx - win + 1 : v_idx + 1]
                    vals = [raw_flows[idx] for idx in window_indices]
                    if filter_type == 'Mean':
                        self.flow_rates[i] = statistics.mean(vals)
                    else:
                        self.flow_rates[i] = statistics.median(vals)
                        
        elif filter_type == 'EMA':
            try: alpha = float(self.ent_filter_param.text())
            except: alpha = 0.1
            alpha = max(0.01, min(1.0, alpha))
            
            self.flow_rates[valid_indices[0]] = raw_flows[valid_indices[0]]
            for v_idx in range(1, len(valid_indices)):
                i = valid_indices[v_idx]
                prev_i = valid_indices[v_idx-1]
                if self.flow_rates[prev_i] is not None:
                    self.flow_rates[i] = alpha * raw_flows[i] + (1 - alpha) * self.flow_rates[prev_i]
                else:
                    self.flow_rates[i] = raw_flows[i]
                    
        elif filter_type == 'Butterworth':
            try: cutoff = float(self.ent_filter_param.text())
            except: cutoff = 0.05
            
            try:
                try: intrvl = float(self.ent_interval.text())
                except: intrvl = 1.0
                fs = 1.0 / max(0.1, intrvl)
                nyq = 0.5 * fs
                normal_cutoff = cutoff / nyq
                
                if len(valid_indices) > 15 and normal_cutoff < 1.0:
                    import scipy.signal
                    b, a = scipy.signal.butter(2, normal_cutoff, btype='low', analog=False)
                    raw_vals = [raw_flows[idx] for idx in valid_indices]
                    filtered_vals = scipy.signal.filtfilt(b, a, raw_vals)
                    for v_idx, val in enumerate(filtered_vals):
                        self.flow_rates[valid_indices[v_idx]] = val
                else:
                    for i in valid_indices:
                        self.flow_rates[i] = raw_flows[i]
            except:
                for i in valid_indices:
                    self.flow_rates[i] = raw_flows[i]
                    
        elif filter_type == 'Adaptive':
            try: rpm = float(self.ent_rpm.text())
            except: rpm = 30.0
            if rpm <= 0: rpm = 1.0
            try: rollers = float(self.ent_rollers.text())
            except: rollers = 3.0
            if rollers <= 0: rollers = 1.0
            
            pulse_hz = (rpm / 60.0) * rollers
            period_sec = 1.0 / pulse_hz
            try: intrvl = float(self.ent_interval.text())
            except: intrvl = 1.0
            win = max(1, int(round(period_sec / max(0.1, intrvl))))
            
            for v_idx in range(len(valid_indices)):
                if v_idx >= win - 1:
                    i = valid_indices[v_idx]
                    window_indices = valid_indices[v_idx - win + 1 : v_idx + 1]
                    vals = [raw_flows[idx] for idx in window_indices]
                    self.flow_rates[i] = statistics.mean(vals)

    def update_gui_components(self):
        n = len(self.times_sec)
        n_items = self.table.rowCount()
        needs_redraw = False
        
        if n > getattr(self, 'last_flow_calc_n', 0):
            self.calculate_flows()
            self.last_flow_calc_n = n
            needs_redraw = True
            
            if self.recording and self.chk_auto_stop.isChecked():
                if len(self.flow_rates) > 0 and self.flow_rates[-1] is not None:
                    current_flow = self.flow_rates[-1]
                    try: thresh = float(self.ent_auto_stop_thresh.text())
                    except: thresh = 0.1
                    
                    if abs(current_flow) > thresh:
                        self.last_activity_time = time.time()
                    else:
                        try: stop_min = float(self.ent_auto_stop_min.text())
                        except: stop_min = 5.0
                        if time.time() - getattr(self, 'last_activity_time', time.time()) > (stop_min * 60.0):
                            self.toggle_record()
                            self.txt_notes.append(f"\n[Auto-Stop] Triggered at {datetime.datetime.now().strftime('%H:%M:%S')} due to flow rate < {thresh} for {stop_min} min.")
                            
                            if self.combo_autosave.currentText() == "On Auto-Stop":
                                self.save_excel(auto=True)
                                self.txt_notes.append(f"[Auto-Save] Data saved to Data folder automatically.")
            
            if self.recording and self.combo_autosave.currentText() == "Every X Minutes":
                try: autosave_interval = float(self.ent_autosave_min.text())
                except: autosave_interval = 5.0
                
                if len(self.times_sec) > 0:
                    elapsed_min = self.times_sec[-1] / 60.0
                    last_autosave = getattr(self, 'last_autosave_min', 0.0)
                    if (elapsed_min - last_autosave) >= autosave_interval:
                        self.save_excel(auto=True)
                        self.last_autosave_min = elapsed_min
        if needs_redraw or n > n_items:
            try:
                # Update plot
                f_v = []
                if len(self.times_min) > 0:
                    if self.combo_xaxis.currentText() == "Timestamp":
                        x_data = self.get_wall_clock_data()
                        self.ax_mass.set_xlabel("Elapsed Timestamp (min)")
                    else:
                        x_data = self.times_min
                        self.ax_mass.set_xlabel("Duration (min)")
                        
                    self.line_mass.set_data(list(x_data), list(self.weights))
                    
                    f_v = [f for f in self.flow_rates if f is not None]
                    if self.chk_show_flow.isChecked() and len(f_v) > 0:
                        t_v = [x_data[i] for i in range(len(x_data)) if self.flow_rates[i] is not None]
                        self.line_flow.set_data(list(t_v), list(f_v))
                
                if self.unit_set:
                    self.ax_mass.set_ylabel(f"Weight ({self.current_unit})")
                    
                if self.chk_auto_x.isChecked() or self.chk_auto_y_mass.isChecked() or self.chk_auto_y_flow.isChecked():
                    self.ax_mass.relim()
                    if self.chk_auto_x.isChecked(): self.ax_mass.autoscale(enable=True, axis='x', tight=None)
                    if self.chk_auto_y_mass.isChecked(): self.ax_mass.autoscale(enable=True, axis='y', tight=None)
                    if self.chk_show_flow.isChecked() and len(f_v) > 0:
                        self.ax_flow.relim()
                        if self.chk_auto_y_flow.isChecked(): self.ax_flow.autoscale(enable=True, axis='y', tight=None)
                        
                    xl = self.ax_mass.get_xlim()
                    yl = self.ax_mass.get_ylim()
                    if self.chk_auto_x.isChecked():
                        self.ent_xmin.setText(str(round(xl[0], 2)))
                        self.ent_xmax.setText(str(round(xl[1], 2)))
                    if self.chk_auto_y_mass.isChecked():
                        self.ent_ymin.setText(str(round(yl[0], 2)))
                        self.ent_ymax.setText(str(round(yl[1], 2)))
                    
                    if self.chk_show_flow.isChecked() and len(f_v) > 0:
                        fl = self.ax_flow.get_ylim()
                        if self.chk_auto_y_flow.isChecked():
                            self.ent_flowmin.setText(str(round(fl[0], 2)))
                            self.ent_flowmax.setText(str(round(fl[1], 2)))
                            
                self.canvas.draw_idle()
                
                # Update Table
                start_idx = n_items
                self.table.setRowCount(n)
                for i in range(start_idx, n):
                    if i < len(self.flow_rates) and self.flow_rates[i] is not None:
                        flow_str = f"{self.flow_rates[i]:.4f}"
                    else: 
                        flow_str = ""
                    self.table.setItem(i, 0, QTableWidgetItem(self.timestamps[i].split(" ")[1]))
                    self.table.setItem(i, 1, QTableWidgetItem(f"{self.times_min[i]:.3f}"))
                    self.table.setItem(i, 2, QTableWidgetItem(f"{self.weights[i]:.3f}"))
                    self.table.setItem(i, 3, QTableWidgetItem(flow_str))
                
                if n > 0:
                    self.table.scrollToBottom()
            except Exception as e:
                pass

    def get_wall_clock_data(self):
        if not hasattr(self, 'wall_clock_min'):
            self.wall_clock_min = []
            
        if len(self.wall_clock_min) < len(self.timestamps):
            try:
                if len(self.wall_clock_min) == 0:
                    self._t0 = datetime.datetime.strptime(self.timestamps[0], "%Y-%m-%d %H:%M:%S")
                for i in range(len(self.wall_clock_min), len(self.timestamps)):
                    t = datetime.datetime.strptime(self.timestamps[i], "%Y-%m-%d %H:%M:%S")
                    self.wall_clock_min.append((t - self._t0).total_seconds() / 60.0)
            except:
                while len(self.wall_clock_min) < len(self.timestamps):
                    self.wall_clock_min.append(self.times_min[len(self.wall_clock_min)])
        return self.wall_clock_min

    def toggle_record(self):
        if not self.recording:
            if len(self.times_sec) == 0:
                self.start_time = time.time()
                self.last_autosave_min = 0.0
            if hasattr(self, 'wall_clock_min'):
                self.wall_clock_min.clear()
            self.recording = True
            self.last_activity_time = time.time()
            self.ax_mass.set_title(f"{self.tab_name} - Live Weight Data (RECORDING)", color='black')
            self.btn_record.setStyleSheet("background-color: #e74c3c; color: white;")
            self.update_button_text()
        else:
            self.recording = False
            self.ax_mass.set_title(f"{self.tab_name} - Live Weight Data (PAUSED)", color='black')
            self.btn_record.setStyleSheet("background-color: #2ecc71; color: white; font-weight: bold;")
            self.update_button_text()
        self.canvas.draw_idle()

    def tare_balance(self):
        if self.serial_thread:
            self.serial_thread.send_tare()

    def clear_data(self):
        self.times_sec.clear()
        self.times_min.clear()
        self.weights.clear()
        self.timestamps.clear()
        self.flow_rates.clear()
        if hasattr(self, 'wall_clock_min'): self.wall_clock_min.clear()
        self.start_time = time.time()
        self.last_activity_time = time.time()
        self.last_recorded_time = 0.0
        self.last_flow_calc_n = 0
        
        self.line_mass.set_data([], [])
        self.line_flow.set_data([], [])
        self.line_fit.set_data([], [])
        self.canvas.draw_idle()
        
        self.table.setRowCount(0)

    def apply_axis_limits(self):
        if not self.chk_auto_x.isChecked():
            try: self.ax_mass.set_xlim(float(self.ent_xmin.text()), float(self.ent_xmax.text()))
            except: pass
        if not self.chk_auto_y_mass.isChecked():
            try: self.ax_mass.set_ylim(float(self.ent_ymin.text()), float(self.ent_ymax.text()))
            except: pass
        if not self.chk_auto_y_flow.isChecked() and self.chk_show_flow.isChecked():
            try: self.ax_flow.set_ylim(float(self.ent_flowmin.text()), float(self.ent_flowmax.text()))
            except: pass
        self.canvas.draw_idle()

    def manual_apply_filters(self):
        self.last_flow_calc_n = -1
        self.update_gui_components()

    def apply_linear_fit(self):
        if len(self.times_min) < 2: return
        try:
            start_min = float(self.ent_fit_start.text())
            end_min = float(self.ent_fit_end.text())
        except: return
            
        valid_indices = [i for i, t in enumerate(self.times_min) if start_min <= t <= end_min]
        if len(valid_indices) < 2: return
            
        x_vals = [self.times_min[i] for i in valid_indices]
        y_vals = [self.weights[i] for i in valid_indices]
        
        try:
            slope, intercept, r_value, p_value, std_err = scipy.stats.linregress(x_vals, y_vals)
            fit_y = [slope * x + intercept for x in x_vals]
            self.line_fit.set_data(x_vals, fit_y)
            self.canvas.draw_idle()
            
            note = f"\n[Linear Fit: {start_min} to {end_min} min] Avg Flow Rate: {slope:.4f} g/min | R²: {r_value**2:.4f}"
            self.txt_notes.append(note)
            self.unsaved_changes = True
            if self.app: self.app.set_unsaved_state(self.tab_name, True)
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def prune_backup_file(self):
        if not os.path.exists(self.backup_path): return
        try:
            with open(self.backup_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                rows = list(reader)
                
            if not header: return
            valid_rows = []
            thirty_days_ago = datetime.datetime.now() - datetime.timedelta(days=30)
            
            for row in rows:
                if len(row) > 0:
                    try:
                        ts_str = row[0]
                        if len(ts_str) <= 19: ts = datetime.datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S")
                        else: ts = datetime.datetime.strptime(ts_str, "%Y-%m-%d %H:%M:%S.%f")
                        if ts >= thirty_days_ago: valid_rows.append(row)
                    except: valid_rows.append(row)
                        
            with open(self.backup_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(header)
                writer.writerows(valid_rows)
        except: pass

    def log_calibration(self):
        try: rpm = float(self.ent_rpm.text())
        except: return QMessageBox.critical(self, "Error", "Invalid RPM.")
        if len(self.flow_rates) == 0 or len(self.times_sec) == 0: return
            
        current_time = self.times_sec[-1]
        stable_flows = []
        for i in range(len(self.times_sec)-1, -1, -1):
            if self.flow_rates[i] is not None:
                if current_time - self.times_sec[i] <= 30.0: stable_flows.append(self.flow_rates[i])
                else: break
                    
        if not stable_flows: return
        flow = sum(stable_flows) / len(stable_flows)
        
        cal_path = os.path.join(os.getcwd(), "Data Backups", f"{self.tab_name.replace(' ', '_')}_Calibration.csv")
        is_new = not os.path.exists(cal_path)
        
        try:
            with open(cal_path, 'a', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                if is_new: writer.writerow(["Timestamp", "Pump RPM", "Avg Flow Rate (g/min)", "Averaging Window (s)"])
                ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                writer.writerow([ts, rpm, round(flow, 4), "30.0"])
            QMessageBox.information(self, "Success", f"Logged: {round(flow,4)} g/min")
        except: pass

    def browse_save_folder(self):
        folder = QFileDialog.getExistingDirectory(self, "Select Save Folder", self.ent_save_folder.text())
        if folder:
            self.ent_save_folder.setText(folder)

    def toggle_autosave_interval(self):
        self.ent_autosave_min.setEnabled(self.combo_autosave.currentText() == "Every X Minutes")

    def toggle_template_source(self):
        use_global = self.chk_use_global_template.isChecked()
        self.ent_save_template.setEnabled(not use_global)
        if use_global and self.app:
            self.ent_save_template.setText(self.app.config.get("global_filename_template", "<Tab Name>_<Date>_<Time>"))

    def get_resolved_string(self, template_str):
        if not template_str: return ""
        s = template_str
        s = s.replace("<Tab Name>", self.tab_name)
        s = s.replace("<Date>", datetime.datetime.now().strftime("%Y-%m-%d"))
        s = s.replace("<Time>", datetime.datetime.now().strftime("%H%M%S"))
        global_tokens = self.app.config.get("global_tokens", []) if self.app else []
        for t in global_tokens:
            name = t.get("name", "").strip()
            if name:
                if hasattr(self, 'token_inputs') and name in self.token_inputs:
                    val = self.token_inputs[name].text()
                else:
                    val = t.get("value", "")
                s = s.replace(f"<{name}>", val)
        return s

    def save_excel(self, auto=False):
        if not self.times_sec: return
        
        if self.chk_use_global_template.isChecked():
            template = self.app.config.get("global_filename_template", "<Tab Name>_<Date>_<Time>") if self.app else "<Tab Name>_<Date>_<Time>"
        else:
            template = self.ent_save_template.text()
            if not template.strip(): template = "<Tab Name>_<Date>_<Time>"
            
        filename = self.get_resolved_string(template)
        if not filename.endswith(".xlsx"): filename += ".xlsx"
        filename = filename.replace(' ', '_')
        
        save_dir = self.ent_save_folder.text().strip()
        if not save_dir: save_dir = "Data"
        
        os.makedirs(save_dir, exist_ok=True)
        default_path = os.path.join(save_dir, filename)
        
        if auto:
            filepath = default_path
        else:
            filepath, _ = QFileDialog.getSaveFileName(self, "Save Excel Data", default_path, "Excel files (*.xlsx)")
            if not filepath: return
            
        try:
            wb = openpyxl.Workbook()
            doc_ws = wb.active
            doc_ws.title = "Documentation"
            
            doc_ws["A1"] = f"{self.tab_name} - Experiment Documentation"
            doc_ws["A1"].font = Font(size=14, bold=True)
            doc_ws["A3"] = "Export Date:"
            doc_ws["B3"] = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            doc_ws["A4"] = "Balance Brand:"
            doc_ws["B4"] = self.combo_brand.currentText()
            doc_ws["A5"] = "COM Port:"
            doc_ws["B5"] = self.combo_com.currentText()
            
            current_row = 7
            global_tokens = self.app.config.get("global_tokens", []) if self.app else []
            for t in global_tokens:
                n = t.get("name", "").strip()
                if n:
                    doc_ws[f"A{current_row}"] = f"{n}:"
                    doc_ws[f"B{current_row}"] = t.get("value", "")
                    current_row += 1
            
            current_row += 1
            doc_ws[f"A{current_row}"] = "Experiment Notes:"
            doc_ws[f"A{current_row}"].font = Font(bold=True)
            doc_ws[f"A{current_row+1}"] = self.get_resolved_string(self.txt_notes.toPlainText())
            doc_ws.column_dimensions["A"].width = 25
            doc_ws.column_dimensions["B"].width = 30
            
            data_ws = wb.create_sheet("Data")
            data_ws.append(["Timestamp", "Duration (min)", "Mass (g)", "Flow Rate (g/min)"])
            for cell in data_ws[1]: cell.font = Font(bold=True)
                
            for i in range(len(self.times_sec)):
                ts = self.timestamps[i]
                dur = round(self.times_min[i], 5)
                mass = self.weights[i]
                flw = self.flow_rates[i] if i < len(self.flow_rates) and self.flow_rates[i] is not None else ""
                data_ws.append([ts, dur, mass, flw])
                
            chart_ws = wb.create_sheet("Chart")
            chart = ScatterChart()
            chart.title = f"{self.tab_name} - Weight vs. Time"
            max_row = len(self.times_sec) + 1
            if max_row > 1:
                xvalues = Reference(data_ws, min_col=2, min_row=2, max_row=max_row)
                yvalues = Reference(data_ws, min_col=3, min_row=2, max_row=max_row)
                series = Series(yvalues, xvalues, title="Recorded Weight")
                chart.series.append(series)
            chart_ws.add_chart(chart, "A1")
                
            wb.save(filepath)
            self.unsaved_changes = False
            if self.app: self.app.set_unsaved_state(self.tab_name, False)
            QMessageBox.information(self, "Success", "Saved to Excel.")
        except Exception as e:
            QMessageBox.critical(self, "Error", str(e))

    def save_graph(self):
        dt_str = datetime.datetime.now().strftime("%Y-%m-%d_%H%M")
        filepath, _ = QFileDialog.getSaveFileName(self, "Save Graph", f"{self.tab_name.replace(' ', '_')}_graph_{dt_str}.png", "PNG Image (*.png)")
        if filepath:
            try:
                self.fig.savefig(filepath)
                QMessageBox.information(self, "Success", "Graph saved.")
            except Exception as e:
                QMessageBox.critical(self, "Error", str(e))

    def scan_backup_sessions(self):
        if not os.path.exists(self.backup_path): return []
        sessions = []
        try:
            with open(self.backup_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                header = next(reader, None)
                rows = list(reader)
            current_session = None
            last_dur = -1.0
            for idx, row in enumerate(rows):
                if len(row) >= 4:
                    try:
                        ts, dur = row[0], float(row[1])
                        if current_session is None or dur < last_dur:
                            if current_session:
                                current_session["end_idx"] = idx - 1
                                sessions.append(current_session)
                            current_session = {"start_idx": idx, "start_ts": ts, "end_ts": ts, "points": 0}
                        current_session["end_ts"] = ts
                        current_session["points"] += 1
                        last_dur = dur
                    except: pass
            if current_session:
                current_session["end_idx"] = len(rows) - 1
                sessions.append(current_session)
        except: pass
        return sessions

    def prompt_session_recovery(self):
        sessions = self.scan_backup_sessions()
        if not sessions: return QMessageBox.information(self, "Recovery", "No recovery sessions found.")
        
        dialog = QDialog(self)
        dialog.setWindowTitle("Restore Session")
        dialog.setMinimumSize(400, 300)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Multiple recording sessions were found. Please select which one to restore:"))
        
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        
        bg = QButtonGroup(dialog)
        for i, s in enumerate(sessions):
            rb = QRadioButton(f"Session {i+1} | {s['start_ts']} -> {s['end_ts']} | {s['points']} pts")
            if i == len(sessions)-1: rb.setChecked(True)
            bg.addButton(rb, i)
            scroll_layout.addWidget(rb)
        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
        btn_layout = QHBoxLayout()
        btn_restore = QPushButton("Restore Selected")
        btn_ignore = QPushButton("Ignore")
        btn_layout.addWidget(btn_restore)
        btn_layout.addWidget(btn_ignore)
        layout.addLayout(btn_layout)
        
        def do_restore():
            idx = bg.checkedId()
            if idx >= 0:
                s = sessions[idx]
                try:
                    with open(self.backup_path, 'r', newline='', encoding='utf-8') as f:
                        reader = csv.reader(f)
                        header = next(reader, None)
                        rows = list(reader)
                    for r_idx in range(s["start_idx"], s["end_idx"] + 1):
                        row = rows[r_idx]
                        if len(row) >= 4:
                            try:
                                self.timestamps.append(row[0])
                                dur = float(row[1])
                                self.times_min.append(dur)
                                self.times_sec.append(dur * 60.0)
                                self.weights.append(float(row[2]))
                                flw = row[3]
                                if flw: self.flow_rates.append(float(flw))
                            except: pass
                    if self.times_sec:
                        self.start_time = time.time() - self.times_sec[-1]
                        self.last_recorded_time = self.times_sec[-1]
                        self.unsaved_changes = True
                except: pass
            dialog.accept()
            
        btn_restore.clicked.connect(do_restore)
        btn_ignore.clicked.connect(dialog.reject)
        dialog.exec()

    def save_tab_settings(self):
        if not self.app: return
        kb = self.app.config.get("known_balances", {})
        if self.tab_name not in kb:
            kb[self.tab_name] = {"brand": self.combo_brand.currentText(), "port": self.combo_com.currentText(), "unsaved": False}
        kb[self.tab_name]["settings"] = {
            "interval": self.ent_interval.text(),
            "savgol_win": self.ent_savgol_win.text(),
            "savgol_poly": self.ent_savgol_poly.text(),
            "filter_type": self.combo_filter.currentText(),
            "filter_param": self.ent_filter_param.text(),
            "rpm": self.ent_rpm.text(),
            "rollers": self.ent_rollers.text(),
            "do_auto_scale_x": self.chk_auto_x.isChecked(),
            "do_auto_scale_y_mass": self.chk_auto_y_mass.isChecked(),
            "do_auto_scale_y_flow": self.chk_auto_y_flow.isChecked(),
            "do_auto_stop": self.chk_auto_stop.isChecked(),
            "auto_stop_min": self.ent_auto_stop_min.text(),
            "auto_stop_thresh": self.ent_auto_stop_thresh.text(),
            "experiment_notes": self.txt_notes.toPlainText(),
            "save_folder": self.ent_save_folder.text(),
            "use_global_template": self.chk_use_global_template.isChecked(),
            "save_template": self.ent_save_template.text(),
            "autosave_mode": self.combo_autosave.currentText(),
            "autosave_min": self.ent_autosave_min.text(),
            "xmin": self.ent_xmin.text(),
            "xmax": self.ent_xmax.text(),
            "ymin": self.ent_ymin.text(),
            "ymax": self.ent_ymax.text(),
            "flowmin": self.ent_flowmin.text(),
            "flowmax": self.ent_flowmax.text()
        }
        
        if hasattr(self, 'token_inputs'):
            kb[self.tab_name]["settings"]["tab_tokens"] = {k: v.text() for k, v in self.token_inputs.items()}

    def load_tab_settings(self):
        if not self.app: return
        kb = self.app.config.get("known_balances", {})
        if self.tab_name in kb and "settings" in kb[self.tab_name]:
            settings = kb[self.tab_name]["settings"]
            if "interval" in settings: self.ent_interval.setText(settings["interval"])
            if "savgol_win" in settings: self.ent_savgol_win.setText(settings["savgol_win"])
            if "savgol_poly" in settings: self.ent_savgol_poly.setText(settings["savgol_poly"])
            if "filter_type" in settings: self.combo_filter.setCurrentText(settings["filter_type"])
            if "filter_param" in settings: self.ent_filter_param.setText(settings["filter_param"])
            if "rpm" in settings: self.ent_rpm.setText(settings["rpm"])
            if "rollers" in settings: self.ent_rollers.setText(settings["rollers"])
            if "do_auto_scale_x" in settings: self.chk_auto_x.setChecked(settings["do_auto_scale_x"])
            if "do_auto_scale_y_mass" in settings: self.chk_auto_y_mass.setChecked(settings["do_auto_scale_y_mass"])
            if "do_auto_scale_y_flow" in settings: self.chk_auto_y_flow.setChecked(settings["do_auto_scale_y_flow"])
            if "do_auto_stop" in settings: self.chk_auto_stop.setChecked(settings["do_auto_stop"])
            if "auto_stop_min" in settings: self.ent_auto_stop_min.setText(settings["auto_stop_min"])
            if "auto_stop_thresh" in settings: self.ent_auto_stop_thresh.setText(settings["auto_stop_thresh"])
            if "experiment_notes" in settings: self.txt_notes.setText(settings["experiment_notes"])
            if "save_folder" in settings: self.ent_save_folder.setText(settings["save_folder"])
            
            if "xmin" in settings: self.ent_xmin.setText(settings["xmin"])
            if "xmax" in settings: self.ent_xmax.setText(settings["xmax"])
            if "ymin" in settings: self.ent_ymin.setText(settings["ymin"])
            if "ymax" in settings: self.ent_ymax.setText(settings["ymax"])
            if "flowmin" in settings: self.ent_flowmin.setText(settings["flowmin"])
            if "flowmax" in settings: self.ent_flowmax.setText(settings["flowmax"])
            
            if "use_global_template" in settings:
                self.chk_use_global_template.setChecked(settings["use_global_template"])
            if "save_template" in settings:
                self.ent_save_template.setText(settings["save_template"])
                
            if "autosave_mode" in settings: self.combo_autosave.setCurrentText(settings["autosave_mode"])
            if "autosave_min" in settings: self.ent_autosave_min.setText(settings["autosave_min"])
            
            if "tab_tokens" in settings and hasattr(self, 'token_inputs'):
                for name, val in settings["tab_tokens"].items():
                    if name in self.token_inputs:
                        self.token_inputs[name].setText(val)

    def save_tab_config(self):
        settings_dict = {
            "interval": self.ent_interval.text(),
            "savgol_win": self.ent_savgol_win.text(),
            "savgol_poly": self.ent_savgol_poly.text(),
            "filter_type": self.combo_filter.currentText(),
            "filter_param": self.ent_filter_param.text(),
            "rpm": self.ent_rpm.text(),
            "rollers": self.ent_rollers.text(),
            "do_auto_scale_x": self.chk_auto_x.isChecked(),
            "do_auto_scale_y_mass": self.chk_auto_y_mass.isChecked(),
            "do_auto_scale_y_flow": self.chk_auto_y_flow.isChecked(),
            "do_auto_stop": self.chk_auto_stop.isChecked(),
            "auto_stop_min": self.ent_auto_stop_min.text(),
            "auto_stop_thresh": self.ent_auto_stop_thresh.text(),
            "experiment_notes": self.txt_notes.toPlainText(),
            "save_folder": self.ent_save_folder.text(),
            "use_global_template": self.chk_use_global_template.isChecked(),
            "save_template": self.ent_save_template.text(),
            "autosave_mode": self.combo_autosave.currentText(),
            "autosave_min": self.ent_autosave_min.text(),
            "xmin": self.ent_xmin.text(),
            "xmax": self.ent_xmax.text(),
            "ymin": self.ent_ymin.text(),
            "ymax": self.ent_ymax.text(),
            "flowmin": self.ent_flowmin.text(),
            "flowmax": self.ent_flowmax.text()
        }
        
        if hasattr(self, 'token_inputs'):
            kb[self.tab_name]["settings"]["tab_tokens"] = {k: v.text() for k, v in self.token_inputs.items()}
        if hasattr(self, 'token_inputs'):
            settings_dict["tab_tokens"] = {k: v.text() for k, v in self.token_inputs.items()}
            
        cfg = {
            "name": self.tab_name,
            "brand": self.combo_brand.currentText(),
            "port": self.combo_com.currentText(),
            "timestamp": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "settings": settings_dict
        }
        saved = self.app.config.setdefault("saved_tabs", [])
        found = False
        for i, s in enumerate(saved):
            if s.get("name") == self.tab_name:
                saved[i] = cfg
                found = True
                break
        if not found:
            saved.append(cfg)
            
        self.app.save_config()
        QMessageBox.information(self, "Success", "Tab Configuration Saved.")

    def load_tab_config_dialog(self):
        saved_tabs = self.app.config.get("saved_tabs", [])
        if not saved_tabs:
            QMessageBox.information(self, "Info", "No saved tab configurations found.")
            return
            
        dialog = QDialog(self)
        dialog.setWindowTitle("Load Tab Configuration")
        dialog.setMinimumSize(400, 300)
        layout = QVBoxLayout(dialog)
        layout.addWidget(QLabel("Select a saved tab configuration to load as a new tab:"))
        
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout(scroll_widget)
        bg = QButtonGroup(dialog)
        
        for i, cfg in enumerate(saved_tabs):
            text = f"{cfg.get('name', 'Unknown')} - {cfg.get('brand', '')} on {cfg.get('port', '')} ({cfg.get('timestamp', '')})"
            rb = QRadioButton(text)
            if i == 0: rb.setChecked(True)
            bg.addButton(rb, i)
            scroll_layout.addWidget(rb)
            
        scroll_widget.setLayout(scroll_layout)
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
        btn_layout = QHBoxLayout()
        btn_load = QPushButton("Load Selected")
        btn_delete = QPushButton("Delete Selected")
        btn_delete.setStyleSheet("background-color: #e74c3c; color: white;")
        
        btn_layout.addWidget(btn_load)
        btn_layout.addWidget(btn_delete)
        layout.addLayout(btn_layout)
        
        def do_load():
            idx = bg.checkedId()
            if idx >= 0:
                cfg = saved_tabs[idx]
                if "settings" in cfg:
                    kb = self.app.config.setdefault("known_balances", {})
                    kb[cfg.get("name", "New Tab")] = {
                        "brand": cfg.get("brand", ""),
                        "port": cfg.get("port", ""),
                        "settings": cfg["settings"],
                        "unsaved": False
                    }
                self.app.add_tab_with_settings(cfg.get("name", "New Tab"), cfg.get("brand", ""), cfg.get("port", ""))
            dialog.accept()
            
        def do_delete():
            idx = bg.checkedId()
            if idx >= 0:
                del saved_tabs[idx]
                self.app.save_config()
                dialog.accept()
                self.load_tab_config_dialog()
            
        btn_load.clicked.connect(do_load)
        btn_delete.clicked.connect(do_delete)
        btn_cancel.clicked.connect(dialog.reject)
        dialog.exec()

    def close_tab(self):
        if self.recording:
            QMessageBox.warning(self, "Warning", "Stop recording before closing.")
            return
        if self.app:
            self.app.close_tab(self)

    def request_add_tab(self):
        if self.app:
            self.app.add_tab(self.ent_new_tab.text())
            self.ent_new_tab.clear()

    def request_rename_tab(self):
        if self.app:
            self.app.rename_tab()


class MultiBalanceApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Multi-Balance Data Streamer (PyQt6)")
        self.resize(1400, 800)
        
        self.config_path = os.path.join(os.getcwd(), "config.json")
        self.config = {"auto_connect": True, "known_balances": {}, "saved_tabs": []}
        if os.path.exists(self.config_path):
            try:
                with open(self.config_path, "r") as f:
                    self.config.update(json.load(f))
            except: pass
            
        # Main Layout
        central = QWidget()
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        
        self.tabs = QTabWidget()
        self.tabs.setMovable(True)
        layout.addWidget(self.tabs)
        
        self.tab_objects = []
        
        has_known = False
        if self.config.get("auto_connect", True):
            for name, details in self.config["known_balances"].items():
                self.add_tab_with_settings(name, details["brand"], details["port"])
                has_known = True
                
        if not has_known:
            self.add_tab()

    def toggle_button_text(self, checked):
        self.config["show_button_text"] = checked
        self.save_config()
        for tab in self.tab_objects:
            if hasattr(tab, "update_button_text"):
                tab.update_button_text()

    def toggle_theme(self, checked=None):
        if checked is not None:
            self.config["dark_mode"] = checked
        else:
            self.config["dark_mode"] = self.action_dark_mode.isChecked()
        self.save_config()
        self.apply_theme()
        
    def apply_theme(self):
        theme = "dark" if self.config.get("dark_mode", True) else "light"
        QApplication.instance().setStyleSheet(qdarktheme.load_stylesheet(theme))

    def show_help_dialog(self):
        msg = QMessageBox(self)
        msg.setWindowTitle("About Flow Rate Engine")
        msg.setText(
            "<b>Derivative Engine:</b><br>"
            "This application uses a Savitzky-Golay (SavGol) filter to compute the flow rate (derivative) from the raw mass data. "
            "Savitzky-Golay fits a local polynomial to the data points, which smooths out high-frequency noise while calculating a highly accurate derivative.<br><br>"
            "<b>Secondary Smoothing:</b><br>"
            "Because flow rate is a derivative, it can be inherently noisy. Secondary Smoothing applies a final low-pass filter to the calculated flow rate for better visualization.<br>"
            "<ul>"
            "<li><b>Mean/Median:</b> Standard moving window averages.</li>"
            "<li><b>EMA:</b> Exponential Moving Average, gives more weight to recent data.</li>"
            "<li><b>Butterworth:</b> A sophisticated signal processing filter for extremely smooth lines.</li>"
            "<li><b>Adaptive:</b> Automatically adjusts its window size based on your pump RPM and roller count.</li>"
            "</ul>"
        )
        msg.exec()

    def restart_app(self):
        self.save_config()
        for tab in self.tab_objects:
            if hasattr(tab, 'save_tab_settings'): tab.save_tab_settings()
            if tab.serial_thread: tab.disconnect_serial()
        
        import subprocess
        subprocess.Popen([sys.executable] + sys.argv)
        QApplication.quit()

    def save_config(self):
        if "known_balances" in self.config:
            new_known = {}
            for i in range(self.tabs.count()):
                tab = self.tabs.widget(i)
                if tab.tab_name in self.config["known_balances"]:
                    new_known[tab.tab_name] = self.config["known_balances"][tab.tab_name]
            for name, details in self.config["known_balances"].items():
                if name not in new_known:
                    new_known[name] = details
            self.config["known_balances"] = new_known

        settings_tab = SettingsTab(self)
        self.tabs.addTab(settings_tab, "Settings")

        try:
            tmp_path = self.config_path + ".tmp"
            with open(tmp_path, "w", encoding="utf-8") as f:
                json.dump(self.config, f, indent=4)
            os.replace(tmp_path, self.config_path)
        except: pass

    def save_connection(self, name, brand, port):
        if "known_balances" not in self.config: self.config["known_balances"] = {}
        self.config["known_balances"][name] = {"brand": brand, "port": port, "unsaved": False}
        self.save_config()

    def set_unsaved_state(self, name, state):
        if "known_balances" in self.config and name in self.config["known_balances"]:
            self.config["known_balances"][name]["unsaved"] = state
            self.save_config()

    def add_tab(self, name=None):
        if not name: name = f"Balance {self.tabs.count() + 1}"
        self.add_tab_with_settings(name, "Bonvoisin", "")

    def add_tab_with_settings(self, name, brand, port):
        tab = BalanceTab(app=self, tab_name=name)
        tab.combo_brand.setCurrentText(brand)
        if port: tab.combo_com.setCurrentText(port)
        self.tabs.addTab(tab, name)
        self.tab_objects.append(tab)
        
        if self.config.get("auto_connect", True) and port:
            QTimer.singleShot(500, lambda: tab.connect_serial(auto=True))

    def rename_tab(self):
        idx = self.tabs.currentIndex()
        if idx < 0: return
        tab = self.tab_objects[idx]
        from PyQt6.QtWidgets import QInputDialog
        new_name, ok = QInputDialog.getText(self, "Rename Tab", "New Name:", text=tab.tab_name)
        if ok and new_name and new_name != tab.tab_name:
            if new_name in [t.tab_name for t in self.tab_objects]: return
            old_name = tab.tab_name
            tab.tab_name = new_name
            self.tabs.setTabText(idx, new_name)
            tab.ax_mass.set_title(f"{new_name} - Live Weight Data")
            tab.canvas.draw_idle()
            
            old_backup = tab.backup_path
            tab.setup_backup_path()
            if os.path.exists(old_backup):
                try: os.rename(old_backup, tab.backup_path)
                except: pass
                
            if old_name in self.config.get("known_balances", {}):
                self.config["known_balances"][new_name] = self.config["known_balances"].pop(old_name)
                self.save_config()

    def close_tab(self, tab):
        idx = self.tabs.indexOf(tab)
        if idx >= 0:
            if hasattr(tab, 'save_tab_settings'): tab.save_tab_settings()
            if tab.serial_thread: tab.disconnect_serial()
            tab.update_timer.stop()
            self.tabs.removeTab(idx)
            self.tab_objects.remove(tab)
            if tab.tab_name in self.config.get("known_balances", {}):
                del self.config["known_balances"][tab.tab_name]
                self.save_config()
            tab.deleteLater()

    def closeEvent(self, event):
        for tab in self.tab_objects:
            if hasattr(tab, 'save_tab_settings'): tab.save_tab_settings()
            if tab.serial_thread: tab.disconnect_serial()
            tab.update_timer.stop()
        self.save_config()
        event.accept()


if __name__ == "__main__":
    app = QApplication(sys.argv)
    
    window = MultiBalanceApp()
    theme = "dark" if window.config.get("dark_mode", True) else "light"
    app.setStyleSheet(qdarktheme.load_stylesheet(theme))
    
    window.showMaximized()
    
    sys.exit(app.exec())
